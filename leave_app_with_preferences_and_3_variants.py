import streamlit as st
import os
os.environ["SUPABASE_URL"] = st.secrets["SUPABASE_URL"]
os.environ["SUPABASE_SERVICE_KEY"] = st.secrets["SUPABASE_SERVICE_KEY"]
os.environ["SUPABASE_KEY"] = st.secrets["SUPABASE_SERVICE_KEY"]
import sys
from datetime import date, datetime, timedelta, time
from io import BytesIO
import pandas as pd
from zoneinfo import ZoneInfo
from openpyxl import load_workbook
from supabase import create_client, Client
import subprocess
import tempfile
from pathlib import Path

# ── page config ──────────────────────────────────────────────────────────────
st.set_page_config(page_title="Glenfield AICU Rota", layout="wide")

# ── secrets ──────────────────────────────────────────────────────────────────
SUPABASE_URL      = st.secrets.get("SUPABASE_URL", "")
SUPABASE_ANON_KEY = st.secrets.get("SUPABASE_ANON_KEY", "")
ALLOWED_EMAIL_DOMAIN = st.secrets.get("ALLOWED_EMAIL_DOMAIN", "")

if not SUPABASE_URL or not SUPABASE_ANON_KEY:
    st.error("Missing SUPABASE_URL or SUPABASE_ANON_KEY in Streamlit secrets.")
    st.stop()

LEAVE_TYPES = ["Annual", "Study", "NOC"]
SHIFT_TYPES = ["A", "B", "D"]
PREF_KINDS  = ["Specific date", "Date range", "Week", "Weekend"]

# ── session state defaults ────────────────────────────────────────────────────
for _k, _v in [("sb_session", None), ("draft_results", []),
               ("draft_done", False), ("draft_period_name", ""),
               ("forgot_password_mode", False)]:
    if _k not in st.session_state:
        st.session_state[_k] = _v

# ── supabase helpers ──────────────────────────────────────────────────────────
def base_client() -> Client:
    return create_client(SUPABASE_URL, SUPABASE_ANON_KEY)

def authed_client(token: str) -> Client:
    c = base_client()
    c.postgrest.auth(token)
    return c

# ── date helpers ──────────────────────────────────────────────────────────────
def validate_dates(s: date, e: date):
    return "End date cannot be before start date." if e < s else None

def overlap(a0, a1, b0, b1):
    return a0 <= b1 and b0 <= a1

def week_bounds(d: date):
    start = d - timedelta(days=d.isoweekday() - 1)
    return start, start + timedelta(days=6)

def weekend_bounds(d: date):
    ws, _ = week_bounds(d)
    return ws + timedelta(days=5), ws + timedelta(days=6)

# ── email confirmation / password recovery handler (runs before auth wall) ────
qp = st.query_params
if "access_token" in qp and "refresh_token" in qp:
    token_type = qp.get("type", "")
    # If this is a password recovery redirect, send the user to the reset page.
    # Supabase sends type=recovery for password resets.
    # Also check type=magiclink as some configs use that for recovery.
    if token_type in ("recovery", "magiclink"):
        st.switch_page("pages/reset_password.py")
    else:
        # Normal email confirmation (signup)
        try:
            tmp = base_client()
            sess = tmp.auth.set_session(qp["access_token"], qp["refresh_token"])
            if sess and sess.session:
                st.session_state["sb_session"] = sess.session.model_dump()
            st.query_params.clear()
            st.success("Email confirmed — you are now signed in.")
            import time as _t; _t.sleep(1)
            st.rerun()
        except Exception:
            st.error("Email confirmation failed. Please sign in manually.")

# ── sidebar: auth ─────────────────────────────────────────────────────────────
with st.sidebar:
    st.title("Glenfield AICU Rota")
    if not st.session_state["sb_session"]:
        # ── Forgot password mode ──────────────────────────────────────────
        if st.session_state.get("forgot_password_mode"):
            st.subheader("Reset your password")
            reset_email = st.text_input("Enter your registered email").strip().lower()
            if st.button("Send reset link", use_container_width=True):
                if not reset_email:
                    st.error("Please enter your email address.")
                elif ALLOWED_EMAIL_DOMAIN and not reset_email.endswith("@" + ALLOWED_EMAIL_DOMAIN):
                    st.error(f"Must be @{ALLOWED_EMAIL_DOMAIN}")
                else:
                    try:
                        c = base_client()
                        c.auth.reset_password_email(reset_email, {
                            "redirect_to": "https://rotaicu.streamlit.app"
                        })
                        st.success("If an account exists for that email, a reset link has been sent. Check your inbox.")
                    except Exception as e:
                        st.error("Failed to send reset email.")
                        st.exception(e)
            if st.button("← Back to sign in", use_container_width=True):
                st.session_state["forgot_password_mode"] = False
                st.rerun()
        else:
            # ── Normal sign in / sign up ──────────────────────────────────
            st.subheader("Sign in / Sign up")
            email    = st.text_input("Email").strip().lower()
            password = st.text_input("Password", type="password")
            if ALLOWED_EMAIL_DOMAIN and email and not email.endswith("@" + ALLOWED_EMAIL_DOMAIN):
                st.warning(f"Must be @{ALLOWED_EMAIL_DOMAIN}")
            c = base_client()
            col1, col2 = st.columns(2)
            with col1:
                if st.button("Sign in", use_container_width=True):
                    if not email or not password:
                        st.error("Enter email and password.")
                    elif ALLOWED_EMAIL_DOMAIN and not email.endswith("@" + ALLOWED_EMAIL_DOMAIN):
                        st.error("Domain not permitted.")
                    else:
                        try:
                            res = c.auth.sign_in_with_password({"email": email, "password": password})
                            st.session_state["sb_session"] = res.session.model_dump() if res.session else None
                            st.rerun()
                        except Exception as e:
                            st.error("Sign-in failed.")
                            st.exception(e)
            with col2:
                if st.button("Sign up", use_container_width=True):
                    if not email or not password:
                        st.error("Enter email and password.")
                    elif ALLOWED_EMAIL_DOMAIN and not email.endswith("@" + ALLOWED_EMAIL_DOMAIN):
                        st.error("Domain not permitted.")
                    else:
                        try:
                            c.auth.sign_up({"email": email, "password": password,
                                            "options": {"email_redirect_to": "https://rotaicu.streamlit.app"}})
                            st.success("Check your email to confirm your account.")
                        except Exception as e:
                            st.error("Sign-up failed.")
                            st.exception(e)
            # ── Forgot password link ──────────────────────────────────────
            if st.button("Forgot your password?", use_container_width=True, type="tertiary"):
                st.session_state["forgot_password_mode"] = True
                st.rerun()

sess = st.session_state["sb_session"]
if not sess:
    st.info("Please sign in using the sidebar to continue.")
    st.stop()

# ── authenticated context ─────────────────────────────────────────────────────
access_token = sess["access_token"]
user_id      = sess["user"]["id"]
user_email   = sess["user"]["email"]
db           = authed_client(access_token)

def is_rota_admin() -> bool:
    try:
        r = db.table("rota_admins").select("user_id").eq("user_id", user_id).execute()
        return bool(r.data)
    except Exception:
        return False

is_admin = is_rota_admin()

with st.sidebar:
    st.markdown("---")
    st.write(f"**{user_email}**")
    st.caption("Rota admin" if is_admin else "Consultant")
    if st.button("Sign out", use_container_width=True):
        st.session_state["sb_session"] = None
        st.rerun()

# ── data fetchers ─────────────────────────────────────────────────────────────
def fetch_periods() -> pd.DataFrame:
    try:
        resp = db.table("rota_periods").select("*").order("start_date").execute()
    except Exception as e:
        st.error("Cannot read rota_periods.")
        st.exception(e)
        return pd.DataFrame()
    df = pd.DataFrame(resp.data or [])
    if not df.empty:
        df["start_date"]   = pd.to_datetime(df["start_date"]).dt.date
        df["end_date"]     = pd.to_datetime(df["end_date"]).dt.date
        df["is_published"] = df["is_published"].astype(bool)
        df["published_at"] = pd.to_datetime(df["published_at"], errors="coerce")
    return df

def fetch_leave() -> pd.DataFrame:
    r = db.table("leave_requests").select("*").order("start_date").execute()
    df = pd.DataFrame(r.data or [])
    if not df.empty:
        df["start_date"] = pd.to_datetime(df["start_date"]).dt.date
        df["end_date"]   = pd.to_datetime(df["end_date"]).dt.date
        df["created_at"] = pd.to_datetime(df["created_at"], errors="coerce")
        df["updated_at"] = pd.to_datetime(df["updated_at"], errors="coerce")
    return df

def fetch_prefs() -> pd.DataFrame:
    r = db.table("preferred_shifts").select("*").order("start_date").execute()
    df = pd.DataFrame(r.data or [])
    if not df.empty:
        df["start_date"] = pd.to_datetime(df["start_date"]).dt.date
        df["end_date"]   = pd.to_datetime(df["end_date"]).dt.date
        df["created_at"] = pd.to_datetime(df["created_at"], errors="coerce")
        df["updated_at"] = pd.to_datetime(df["updated_at"], errors="coerce")
    return df

def publish_due_periods():
    try:
        db.rpc("publish_due_periods", {}).execute()
    except Exception:
        pass

publish_due_periods()
periods = fetch_periods()

def any_published_overlap(s: date, e: date) -> bool:
    if periods.empty:
        return False
    pubs = periods[periods["is_published"]]
    return any(overlap(s, e, r["start_date"], r["end_date"]) for _, r in pubs.iterrows())

def is_period_locked(s: date, e: date) -> bool:
    """True if the date range falls inside a published (locked) period."""
    return (not is_admin) and any_published_overlap(s, e)

# ── countdown banner ──────────────────────────────────────────────────────────
def next_lock_banner():
    if periods.empty or "leave_lock_at" not in periods.columns:
        return
    now = pd.Timestamp.now(tz="UTC")
    df  = periods[~periods["is_published"]].copy()
    df  = df[df["leave_lock_at"].notna()]
    df["_ts"] = pd.to_datetime(df["leave_lock_at"], errors="coerce", utc=True)
    df  = df[df["_ts"].notna() & (df["_ts"] > now)].sort_values("_ts")
    if df.empty:
        return
    row   = df.iloc[0]
    lock  = row["_ts"].to_pydatetime()
    delta = lock - now.to_pydatetime()
    d, h  = delta.days, delta.seconds // 3600
    m     = (delta.seconds % 3600) // 60
    msg   = (f"Leave and shift requests close for **{row.get('name','this period')}** in "
             f"**{d}d {h}h {m}m** ({lock:%Y-%m-%d %H:%M UTC}).")
    st.warning(msg)

# ═════════════════════════════════════════════════════════════════════════════
# MAIN PAGE
# ═════════════════════════════════════════════════════════════════════════════
st.title("Glenfield AICU Rota")
next_lock_banner()

# ── rota periods summary (everyone sees this) ─────────────────────────────────
with st.expander("Rota periods", expanded=False):
    if periods.empty:
        st.info("No rota periods configured yet.")
    else:
        cols = [c for c in ["name","start_date","end_date","is_published","published_at"] if c in periods.columns]
        st.dataframe(periods[cols], hide_index=True, use_container_width=True)
        if periods["is_published"].any():
            st.caption("Requests overlapping a published period are locked.")

# ── period selector (filters what leave/prefs are displayed) ──────────────────
active_period_start = None
active_period_end = None

if not periods.empty:
    periods_sorted = periods.sort_values("start_date", ascending=False).copy()
    periods_sorted["label"] = periods_sorted.apply(
        lambda r: f"{r['name']} ({r['start_date']} → {r['end_date']})", axis=1)
    selected_period_label = st.selectbox(
        "Show leave & preferences for rota period:",
        options=periods_sorted["label"].tolist(),
        key="active_period_filter",
    )
    sel_period_row = periods_sorted[periods_sorted["label"] == selected_period_label].iloc[0]
    active_period_start = sel_period_row["start_date"]
    active_period_end = sel_period_row["end_date"]
    st.caption(f"Displaying requests that overlap **{active_period_start}** to **{active_period_end}**")
else:
    st.info("No rota periods configured — showing all requests.")

def filter_by_period(df: pd.DataFrame) -> pd.DataFrame:
    """Filter a DataFrame with start_date/end_date columns to only rows
    overlapping the active period. Returns all rows if no period is selected."""
    if active_period_start is None or active_period_end is None:
        return df
    if df.empty:
        return df
    return df[df.apply(
        lambda r: overlap(r["start_date"], r["end_date"],
                          active_period_start, active_period_end), axis=1
    )].copy()

# ═════════════════════════════════════════════════════════════════════════════
# SECTION 1 — LEAVE
# ═════════════════════════════════════════════════════════════════════════════
st.subheader("Leave requests")

# ── submit leave ──────────────────────────────────────────────────────────────
with st.expander("Submit a new leave request", expanded=True):
    with st.form("leave_add"):
        c1, c2, c3 = st.columns([2, 1, 1])
        with c1:
            leave_name  = st.text_input("Your name (as on rota)")
            leave_type  = st.selectbox("Leave type", LEAVE_TYPES)
            leave_notes = st.text_input("Notes (optional)")
        with c2:
            leave_start = st.date_input("From", value=date.today(), key="ls")
        with c3:
            leave_end   = st.date_input("To",   value=date.today(), key="le")
        submitted = st.form_submit_button("Submit leave request", use_container_width=True)

    if submitted:
        err = validate_dates(leave_start, leave_end)
        if err:
            st.error(err)
        elif not leave_name.strip():
            st.error("Name is required.")
        elif is_period_locked(leave_start, leave_end):
            st.error("This date range overlaps a published (locked) rota period.")
        else:
            try:
                db.table("leave_requests").insert({
                    "consultant_name": leave_name.strip(),
                    "requester_id":    user_id,
                    "requester_email": user_email,
                    "start_date":      leave_start.isoformat(),
                    "end_date":        leave_end.isoformat(),
                    "leave_type":      leave_type,
                    "approved":        True,   # auto-approved — no admin approval step
                    "notes":           leave_notes.strip(),
                }).execute()
                st.success("Leave request submitted.")
                st.rerun()
            except Exception as e:
                st.error("Failed to submit.")
                st.exception(e)

# ── my leave ──────────────────────────────────────────────────────────────────
leave_all = fetch_leave()
my_leave  = leave_all[leave_all["requester_id"] == user_id].copy() if not leave_all.empty else pd.DataFrame()
my_leave  = filter_by_period(my_leave)

st.markdown("**Your leave requests**")
if my_leave.empty:
    st.info("You have no leave requests.")
else:
    for _, row in my_leave.iterrows():
        locked = is_period_locked(row["start_date"], row["end_date"])
        col_info, col_edit, col_del = st.columns([4, 1, 1])
        with col_info:
            st.write(f"**{row['consultant_name']}** · {row['leave_type']} · "
                     f"{row['start_date']} → {row['end_date']}"
                     + (f" · _{row['notes']}_" if row.get('notes') else "")
                     + (" 🔒" if locked else ""))
        with col_edit:
            if not locked and st.button("Edit", key=f"le_edit_{row['id']}"):
                st.session_state[f"editing_leave_{row['id']}"] = True
        with col_del:
            if not locked and st.button("Delete", key=f"le_del_{row['id']}"):
                db.table("leave_requests").delete().eq("id", row["id"]).execute()
                st.success("Deleted.")
                st.rerun()

        # inline edit form
        if st.session_state.get(f"editing_leave_{row['id']}"):
            with st.form(f"leave_edit_form_{row['id']}"):
                ec1, ec2, ec3 = st.columns([2, 1, 1])
                with ec1:
                    e_name  = st.text_input("Name", value=row["consultant_name"])
                    e_type  = st.selectbox("Leave type", LEAVE_TYPES,
                                           index=LEAVE_TYPES.index(row["leave_type"]) if row["leave_type"] in LEAVE_TYPES else 0)
                    e_notes = st.text_input("Notes", value=row.get("notes") or "")
                with ec2:
                    e_start = st.date_input("From", value=row["start_date"], key=f"es_{row['id']}")
                with ec3:
                    e_end   = st.date_input("To",   value=row["end_date"],   key=f"ee_{row['id']}")
                save_edit = st.form_submit_button("Save changes")
            if save_edit:
                err = validate_dates(e_start, e_end)
                if err:
                    st.error(err)
                else:
                    db.table("leave_requests").update({
                        "consultant_name": e_name.strip(),
                        "start_date":      e_start.isoformat(),
                        "end_date":        e_end.isoformat(),
                        "leave_type":      e_type,
                        "notes":           e_notes.strip(),
                    }).eq("id", row["id"]).execute()
                    del st.session_state[f"editing_leave_{row['id']}"]
                    st.success("Updated.")
                    st.rerun()

# ═════════════════════════════════════════════════════════════════════════════
# SECTION 2 — PREFERRED SHIFTS
# ═════════════════════════════════════════════════════════════════════════════
st.subheader("Preferred shifts")

# ── submit preference ─────────────────────────────────────────────────────────
with st.expander("Submit a preferred shift request", expanded=True):
    st.caption("You may submit one preference per rota period. "
               "Preferences are soft — the solver will try to honour them.")
    with st.form("pref_add"):
        p1, p2, p3 = st.columns([2, 1, 1])
        with p1:
            pref_name   = st.text_input("Your name (as on rota)", key="pn")
            pref_kind   = st.selectbox("Preference type", PREF_KINDS, key="pk")
            pref_shift  = st.selectbox("Shift type", SHIFT_TYPES, key="ps")
            pref_weight = st.slider("Strength (1 = low, 5 = high)", 1, 5, 3, key="pw")
            pref_notes  = st.text_input("Notes (optional)", key="pno")
        with p2:
            ref_date = st.date_input("Reference date", value=date.today(), key="prd")
        with p3:
            pref_end_override = st.date_input("End date (date range only)", value=date.today(), key="ped")
        pref_submitted = st.form_submit_button("Submit preference", use_container_width=True)

    if pref_submitted:
        if pref_kind == "Specific date":
            ps, pe = ref_date, ref_date
        elif pref_kind == "Date range":
            ps, pe = ref_date, pref_end_override
        elif pref_kind == "Week":
            ps, pe = week_bounds(ref_date)
        else:
            ps, pe = weekend_bounds(ref_date)

        err = validate_dates(ps, pe)
        if err:
            st.error(err)
        elif not pref_name.strip():
            st.error("Name is required.")
        elif is_period_locked(ps, pe):
            st.error("This date range overlaps a published (locked) rota period.")
        else:
            try:
                db.table("preferred_shifts").insert({
                    "consultant_name": pref_name.strip(),
                    "requester_id":    user_id,
                    "requester_email": user_email,
                    "start_date":      ps.isoformat(),
                    "end_date":        pe.isoformat(),
                    "pref_kind":       pref_kind,
                    "shift_type":      pref_shift,
                    "weight":          int(pref_weight),
                    "notes":           pref_notes.strip(),
                }).execute()
                st.success("Preference submitted.")
                st.rerun()
            except Exception as e:
                st.error("Failed to submit.")
                st.exception(e)

# ── my preferences ────────────────────────────────────────────────────────────
prefs_all = fetch_prefs()
my_prefs  = prefs_all[prefs_all["requester_id"] == user_id].copy() if not prefs_all.empty else pd.DataFrame()
my_prefs  = filter_by_period(my_prefs)

# conflict detection: same shift_type + overlapping dates, different user
def pref_conflicts(row) -> list:
    if prefs_all.empty:
        return []
    others = prefs_all[
        (prefs_all["requester_id"] != user_id) &
        (prefs_all["shift_type"] == row["shift_type"]) &
        prefs_all.apply(lambda r: overlap(row["start_date"], row["end_date"],
                                          r["start_date"], r["end_date"]), axis=1)
    ]
    return others["consultant_name"].tolist()

st.markdown("**Your preferred shift requests**")
if my_prefs.empty:
    st.info("You have no preferred shift requests.")
else:
    for _, row in my_prefs.iterrows():
        locked    = is_period_locked(row["start_date"], row["end_date"])
        conflicts = pref_conflicts(row)
        col_info, col_edit, col_del = st.columns([4, 1, 1])
        with col_info:
            st.write(f"**{row['consultant_name']}** · {row['shift_type']} ({row['pref_kind']}) · "
                     f"{row['start_date']} → {row['end_date']} · strength {row['weight']}"
                     + (f" · _{row['notes']}_" if row.get('notes') else "")
                     + (" 🔒" if locked else ""))
            if conflicts:
                st.warning(f"⚠️ {', '.join(conflicts)} has requested the same shift type on overlapping dates.")
        with col_edit:
            if not locked and st.button("Edit", key=f"pe_edit_{row['id']}"):
                st.session_state[f"editing_pref_{row['id']}"] = True
        with col_del:
            if not locked and st.button("Delete", key=f"pe_del_{row['id']}"):
                db.table("preferred_shifts").delete().eq("id", row["id"]).execute()
                st.success("Deleted.")
                st.rerun()

        if st.session_state.get(f"editing_pref_{row['id']}"):
            with st.form(f"pref_edit_form_{row['id']}"):
                ep1, ep2, ep3 = st.columns([2, 1, 1])
                with ep1:
                    e_pname  = st.text_input("Name", value=row["consultant_name"])
                    e_pkind  = st.selectbox("Preference type", PREF_KINDS,
                                            index=PREF_KINDS.index(row["pref_kind"]) if row["pref_kind"] in PREF_KINDS else 0)
                    e_pshift = st.selectbox("Shift type", SHIFT_TYPES,
                                            index=SHIFT_TYPES.index(row["shift_type"]) if row["shift_type"] in SHIFT_TYPES else 0)
                    e_pweight = st.slider("Strength", 1, 5, int(row["weight"]))
                    e_pnotes  = st.text_input("Notes", value=row.get("notes") or "")
                with ep2:
                    e_pstart = st.date_input("From", value=row["start_date"], key=f"eps_{row['id']}")
                with ep3:
                    e_pend   = st.date_input("To",   value=row["end_date"],   key=f"epe_{row['id']}")
                save_pref = st.form_submit_button("Save changes")
            if save_pref:
                err = validate_dates(e_pstart, e_pend)
                if err:
                    st.error(err)
                else:
                    db.table("preferred_shifts").update({
                        "consultant_name": e_pname.strip(),
                        "start_date":      e_pstart.isoformat(),
                        "end_date":        e_pend.isoformat(),
                        "pref_kind":       e_pkind,
                        "shift_type":      e_pshift,
                        "weight":          int(e_pweight),
                        "notes":           e_pnotes.strip(),
                    }).eq("id", row["id"]).execute()
                    del st.session_state[f"editing_pref_{row['id']}"]
                    st.success("Updated.")
                    st.rerun()

# ═════════════════════════════════════════════════════════════════════════════
# ADMIN SECTION
# ═════════════════════════════════════════════════════════════════════════════
if not is_admin:
    st.stop()

st.divider()
st.header("Admin panel")

# ── admin: own leave (same edit/delete as consultants) ────────────────────────
st.subheader("Your leave requests")
admin_leave = leave_all[leave_all["requester_id"] == user_id].copy() if not leave_all.empty else pd.DataFrame()
admin_leave = filter_by_period(admin_leave)

with st.expander("Submit leave for yourself", expanded=False):
    with st.form("admin_leave_add"):
        ac1, ac2, ac3 = st.columns([2, 1, 1])
        with ac1:
            al_name  = st.text_input("Your name (as on rota)", key="al_name")
            al_type  = st.selectbox("Leave type", LEAVE_TYPES, key="al_type")
            al_notes = st.text_input("Notes (optional)", key="al_notes")
        with ac2:
            al_start = st.date_input("From", value=date.today(), key="al_start")
        with ac3:
            al_end   = st.date_input("To",   value=date.today(), key="al_end")
        al_submit = st.form_submit_button("Submit", use_container_width=True)
    if al_submit:
        err = validate_dates(al_start, al_end)
        if err:
            st.error(err)
        elif not al_name.strip():
            st.error("Name is required.")
        else:
            try:
                db.table("leave_requests").insert({
                    "consultant_name": al_name.strip(),
                    "requester_id":    user_id,
                    "requester_email": user_email,
                    "start_date":      al_start.isoformat(),
                    "end_date":        al_end.isoformat(),
                    "leave_type":      al_type,
                    "approved":        True,
                    "notes":           al_notes.strip(),
                }).execute()
                st.success("Leave submitted.")
                st.rerun()
            except Exception as e:
                st.error("Failed.")
                st.exception(e)

if admin_leave.empty:
    st.info("No leave requests from you.")
else:
    for _, row in admin_leave.iterrows():
        col_info, col_edit, col_del = st.columns([4, 1, 1])
        with col_info:
            st.write(f"**{row['consultant_name']}** · {row['leave_type']} · "
                     f"{row['start_date']} → {row['end_date']}"
                     + (f" · _{row['notes']}_" if row.get('notes') else ""))
        with col_edit:
            if st.button("Edit", key=f"adm_le_edit_{row['id']}"):
                st.session_state[f"adm_editing_leave_{row['id']}"] = True
        with col_del:
            if st.button("Delete", key=f"adm_le_del_{row['id']}"):
                db.table("leave_requests").delete().eq("id", row["id"]).execute()
                st.success("Deleted.")
                st.rerun()
        if st.session_state.get(f"adm_editing_leave_{row['id']}"):
            with st.form(f"adm_leave_edit_{row['id']}"):
                aec1, aec2, aec3 = st.columns([2, 1, 1])
                with aec1:
                    ae_name  = st.text_input("Name",  value=row["consultant_name"])
                    ae_type  = st.selectbox("Type", LEAVE_TYPES,
                                            index=LEAVE_TYPES.index(row["leave_type"]) if row["leave_type"] in LEAVE_TYPES else 0)
                    ae_notes = st.text_input("Notes", value=row.get("notes") or "")
                with aec2:
                    ae_start = st.date_input("From", value=row["start_date"], key=f"aes_{row['id']}")
                with aec3:
                    ae_end   = st.date_input("To",   value=row["end_date"],   key=f"aee_{row['id']}")
                ae_save = st.form_submit_button("Save")
            if ae_save:
                err = validate_dates(ae_start, ae_end)
                if err:
                    st.error(err)
                else:
                    db.table("leave_requests").update({
                        "consultant_name": ae_name.strip(),
                        "start_date": ae_start.isoformat(),
                        "end_date":   ae_end.isoformat(),
                        "leave_type": ae_type,
                        "notes":      ae_notes.strip(),
                    }).eq("id", row["id"]).execute()
                    del st.session_state[f"adm_editing_leave_{row['id']}"]
                    st.success("Updated.")
                    st.rerun()

# ── admin: all leave ──────────────────────────────────────────────────────────
st.subheader("All leave requests")
leave_all_filtered = filter_by_period(leave_all)
if leave_all_filtered.empty:
    st.info("No leave requests for the selected rota period.")
else:
    show_cols = [c for c in ["consultant_name","requester_email","start_date","end_date",
                              "leave_type","notes","updated_at"] if c in leave_all_filtered.columns]
    st.dataframe(leave_all_filtered[show_cols], hide_index=True, use_container_width=True)

# ── admin: all preferred shifts ───────────────────────────────────────────────
st.subheader("All preferred shift requests")
prefs_all_filtered = filter_by_period(prefs_all)
if prefs_all_filtered.empty:
    st.info("No preferred shift requests for the selected rota period.")
else:
    pref_cols = [c for c in ["consultant_name","requester_email","start_date","end_date",
                              "pref_kind","shift_type","weight","notes","updated_at"] if c in prefs_all_filtered.columns]
    st.dataframe(prefs_all_filtered[pref_cols], hide_index=True, use_container_width=True)

# ── admin: rota period management ────────────────────────────────────────────
st.subheader("Rota period management")

periods_admin = fetch_periods()
if not periods_admin.empty:
    show_p = [c for c in ["name","start_date","end_date","leave_lock_at",
                           "is_published","published_at"] if c in periods_admin.columns]
    st.dataframe(periods_admin[show_p], hide_index=True, use_container_width=True)

with st.expander("Create / update a rota period", expanded=False):
    with st.form("rota_period_upsert"):
        name_p  = st.text_input("Period name (e.g. Nov 2025 – May 2026)")
        start_p = st.date_input("Start date", value=pd.Timestamp.utcnow().date())
        end_p   = st.date_input("End date",   value=(pd.Timestamp.utcnow() + pd.Timedelta(days=180)).date())
        default_lock_date = (pd.Timestamp.utcnow() + pd.Timedelta(days=14)).date()
        lock_date = st.date_input("Leave lock date (Europe/London)", value=default_lock_date)
        lock_time = st.time_input("Leave lock time (Europe/London)", value=time(17, 0))
        lock_local = pd.Timestamp.combine(lock_date, lock_time).to_pydatetime().replace(tzinfo=ZoneInfo("Europe/London"))
        lock_p = pd.Timestamp(lock_local).tz_convert("UTC").to_pydatetime()
        save_p = st.form_submit_button("Save period", use_container_width=True)

    if save_p:
        try:
            payload = {
                "name":          name_p,
                "start_date":    start_p.isoformat(),
                "end_date":      end_p.isoformat(),
                "leave_lock_at": pd.Timestamp(lock_p).tz_convert("UTC").isoformat(),
            }
            try:
                db.table("rota_periods").upsert(payload, on_conflict="name").execute()
            except Exception:
                existing = db.table("rota_periods").select("id").eq("name", name_p).execute()
                if existing.data:
                    db.table("rota_periods").update(payload).eq("id", existing.data[0]["id"]).execute()
                else:
                    db.table("rota_periods").insert(payload).execute()
            publish_due_periods()
            st.success("Saved.")
            st.rerun()
        except Exception as e:
            st.error("Failed to save.")
            st.exception(e)

with st.expander("Finalise a rota period (locks leave immediately)", expanded=False):
    periods_f = fetch_periods()
    if not periods_f.empty and "start_date" in periods_f.columns:
        periods_f = periods_f.sort_values("start_date", ascending=False)
        periods_f["label"] = periods_f.apply(
            lambda r: f"{r['name']} ({r['start_date']} → {r['end_date']}) — "
                      f"{'PUBLISHED' if r['is_published'] else 'unpublished'}", axis=1)
        sel_f = st.selectbox("Select period to finalise", periods_f["label"].tolist(), key="finalise_sel")
        row_f = periods_f[periods_f["label"] == sel_f].iloc[0]
        st.caption("Finalising marks the period published and locks all leave/preferences immediately.")
        if st.button("Finalise now", type="primary"):
            try:
                now_iso = pd.Timestamp.now(tz="UTC").isoformat()
                upd = {"leave_lock_at": now_iso, "is_published": True, "published_at": now_iso}
                if "is_finalized"  in periods_f.columns: upd["is_finalized"]  = True
                if "finalized_at"  in periods_f.columns: upd["finalized_at"]  = now_iso
                db.table("rota_periods").update(upd).eq("id", row_f["id"]).execute()
                st.success("Period finalised. Consultants are now locked out.")
                st.rerun()
            except Exception as e:
                st.error("Failed.")
                st.exception(e)

# ── admin: draft rota ─────────────────────────────────────────────────────────
st.subheader("Draft rota")
st.caption("Select a published period, upload the base workbook, then run the solver to produce 3 candidate rotas.")

periods_d = fetch_periods()
if periods_d.empty:
    st.warning("No rota periods found. Create and publish one first.")
    st.stop()

periods_d = periods_d.sort_values("start_date", ascending=False).copy()
periods_d["label"] = periods_d.apply(
    lambda r: f"{r['name']} ({r['start_date']} → {r['end_date']}) — "
              f"{'PUBLISHED' if r['is_published'] else 'unpublished'}", axis=1)

sel_label   = st.selectbox("Rota period", periods_d["label"].tolist(), key="draft_period")
sel_row     = periods_d[periods_d["label"] == sel_label].iloc[0]
period_name = str(sel_row["name"])
period_start, period_end = sel_row["start_date"], sel_row["end_date"]
period_published = bool(sel_row["is_published"])

st.caption(f"Period window: {period_start} → {period_end}")

if not period_published:
    st.warning("Drafting is disabled — this period is not yet published.")
    st.stop()

template = st.file_uploader("Upload base rota workbook (.xlsx)", type=["xlsx"], key="draft_template")

col_r1, col_r2 = st.columns([1, 3])
with col_r1:
    if st.button("Reset drafting", key="draft_reset"):
        for k in ["draft_results", "draft_done", "draft_period_name", "draft_template"]:
            st.session_state.pop(k, None)
        st.rerun()
with col_r2:
    st.caption("Reset before uploading a new workbook or changing the period.")

solver_script = Path("solve_rota.py")
if not solver_script.exists():
    st.warning("`solve_rota.py` not found in the repository.")
    st.stop()

relax_week_gap        = st.checkbox("Allow fallback: relax 1-week gap constraint", value=True)
relax_no_consec_wknds = st.checkbox("Allow fallback: relax no-consecutive-weekends constraint", value=True)
hard_cardiac          = st.checkbox("Enforce cardiac XOR as a hard constraint", value=False)
force_truncate        = st.checkbox("Truncate partial-overlap requests to period window", value=False)

variant_seeds = [11, 22, 33]

if template is None:
    st.info("Upload the base rota workbook to enable drafting.")
else:
    # ── Pre-populate shifts UI ────────────────────────────────────────────
    st.markdown("**Pre-populate shifts (optional)**")
    st.caption("Specify shifts you want to lock in before the solver runs. "
               "These will count towards the consultant's allocation target.")

    if "pre_allocs" not in st.session_state:
        st.session_state["pre_allocs"] = []

    with st.expander("Add a pre-allocated shift", expanded=False):
        with st.form("pre_alloc_form"):
            pa_c1, pa_c2, pa_c3 = st.columns([2, 1, 1])
            with pa_c1:
                pa_name = st.text_input("Consultant name (as on rota)", key="pa_name")
                pa_block = st.selectbox("Block type", 
                    ["AB1", "AB2", "DMonThu", "DMonTue", "DWedThu", "WeekendAB", "WeekendMixed"],
                    key="pa_block")
            with pa_c2:
                pa_week = st.date_input("Week starting (Monday)", value=period_start, key="pa_week")
            with pa_c3:
                st.write("")  # spacer
            pa_submit = st.form_submit_button("Add pre-allocation")

        if pa_submit and pa_name.strip():
            st.session_state["pre_allocs"].append({
                "consultant_name": pa_name.strip(),
                "week_start": pa_week,
                "block_type": pa_block,
            })
            st.success(f"Added: {pa_name.strip()} → {pa_block} (week of {pa_week})")
            st.rerun()

    if st.session_state["pre_allocs"]:
        st.markdown("**Current pre-allocations:**")
        for idx, pa in enumerate(st.session_state["pre_allocs"]):
            col_pa, col_del = st.columns([4, 1])
            with col_pa:
                st.write(f"• {pa['consultant_name']} → {pa['block_type']} (week of {pa['week_start']})")
            with col_del:
                if st.button("Remove", key=f"rm_pa_{idx}"):
                    st.session_state["pre_allocs"].pop(idx)
                    st.rerun()

    if st.button("Draft 3 rota versions", key="draft_run", type="primary"):
        st.session_state["draft_results"]     = []
        st.session_state["draft_done"]        = False
        st.session_state["draft_period_name"] = period_name

        with st.spinner("Running solver for 3 variants…"):
            # ── filter leave ──────────────────────────────────────────────
            leave_ok = leave_all.copy() if not leave_all.empty else pd.DataFrame()
            if not leave_ok.empty:
                leave_ok = leave_ok[leave_ok.apply(
                    lambda r: overlap(r["start_date"], r["end_date"], period_start, period_end), axis=1)]
                contained = leave_ok.apply(
                    lambda r: r["start_date"] >= period_start and r["end_date"] <= period_end, axis=1)
                partial = leave_ok[~contained]
                if not partial.empty and not force_truncate:
                    st.error("Some leave partially overlaps the period. Enable truncation or fix dates.")
                    st.dataframe(partial[["consultant_name","start_date","end_date"]], hide_index=True)
                    st.stop()
                if not partial.empty and force_truncate:
                    leave_ok.loc[partial.index, "start_date"] = partial["start_date"].apply(lambda d: max(d, period_start))
                    leave_ok.loc[partial.index, "end_date"]   = partial["end_date"].apply(lambda d: min(d, period_end))
                    contained = leave_ok.apply(
                        lambda r: r["start_date"] >= period_start and r["end_date"] <= period_end, axis=1)
                leave_final = leave_ok[contained].copy()
            else:
                leave_final = leave_ok

            # ── filter prefs ──────────────────────────────────────────────
            pref_ok = prefs_all.copy() if not prefs_all.empty else pd.DataFrame()
            if not pref_ok.empty:
                pref_ok = pref_ok[pref_ok.apply(
                    lambda r: overlap(r["start_date"], r["end_date"], period_start, period_end), axis=1)]
                contained_p = pref_ok.apply(
                    lambda r: r["start_date"] >= period_start and r["end_date"] <= period_end, axis=1)
                partial_p = pref_ok[~contained_p]
                if not partial_p.empty and not force_truncate:
                    st.error("Some preferences partially overlap the period. Enable truncation or fix dates.")
                    st.dataframe(partial_p[["consultant_name","start_date","end_date"]], hide_index=True)
                    st.stop()
                if not partial_p.empty and force_truncate:
                    pref_ok.loc[partial_p.index, "start_date"] = partial_p["start_date"].apply(lambda d: max(d, period_start))
                    pref_ok.loc[partial_p.index, "end_date"]   = partial_p["end_date"].apply(lambda d: min(d, period_end))
                    contained_p = pref_ok.apply(
                        lambda r: r["start_date"] >= period_start and r["end_date"] <= period_end, axis=1)
                pref_final = pref_ok[contained_p].copy()
            else:
                pref_final = pref_ok

            # ── build workbook ────────────────────────────────────────────
            wb = load_workbook(BytesIO(template.getvalue()))
            if "Leave" not in wb.sheetnames:
                st.error("Workbook must contain a 'Leave' sheet.")
                st.stop()
            if "preferred_shifts" not in wb.sheetnames:
                wb.create_sheet("preferred_shifts")

            lws = wb["Leave"]
            for rr in range(2, 5000):
                if lws[f"A{rr}"].value in (None, ""):
                    break
                for col in ("A","B","C","D","E"):
                    lws[f"{col}{rr}"].value = None
            rr = 2
            if not leave_final.empty:
                for _, rec in leave_final.sort_values(["start_date","consultant_name"]).iterrows():
                    lws[f"A{rr}"].value = rec["consultant_name"]
                    lws[f"B{rr}"].value = rec["start_date"]
                    lws[f"C{rr}"].value = rec["end_date"]
                    lws[f"D{rr}"].value = rec["leave_type"]
                    lws[f"E{rr}"].value = True
                    rr += 1

            pws = wb["preferred_shifts"]
            pws.delete_rows(1, pws.max_row or 1)
            pws.append(["Name","StartDate","EndDate","PrefKind","ShiftType","Weight","Notes"])
            if not pref_final.empty:
                for _, rec in pref_final.sort_values(["start_date","consultant_name"]).iterrows():
                    pws.append([rec["consultant_name"], rec["start_date"], rec["end_date"],
                                rec.get("pref_kind"), rec.get("shift_type"),
                                int(rec.get("weight", 3)), rec.get("notes")])

            # ── write PreAllocations sheet ─────────────────────────────────
            pre_allocs_data = st.session_state.get("pre_allocs", [])
            if pre_allocs_data:
                if "PreAllocations" not in wb.sheetnames:
                    wb.create_sheet("PreAllocations")
                paws = wb["PreAllocations"]
                paws.delete_rows(1, paws.max_row or 1)
                paws.append(["Name", "WeekStart", "Block"])
                for pa in pre_allocs_data:
                    paws.append([pa["consultant_name"], pa["week_start"], pa["block_type"]])

            # ── run solver ────────────────────────────────────────────────
            attempts_base = [("Strict", [])]
            if relax_week_gap:
                attempts_base.append(("Relax week-gap", ["--no_hard_week_gap"]))
            if relax_no_consec_wknds:
                attempts_base.append(("Relax no-consec-weekends", ["--no_hard_no_consec_weekends"]))
            if relax_week_gap and relax_no_consec_wknds:
                attempts_base.append(("Relax both", ["--no_hard_week_gap","--no_hard_no_consec_weekends"]))

            results_local = []
            with tempfile.TemporaryDirectory() as td:
                td_path = Path(td)
                solver_input = td_path / "Rota_Master_WITH_Leave.xlsx"
                wb.save(solver_input)
                PYTHON = sys.executable

                for i, seed in enumerate(variant_seeds, start=1):
                    variant_name = f"Variant {i} (seed {seed})"
                    out_file     = td_path / f"Rota_Solved_V{i}.xlsx"
                    succeeded    = False
                    used_label   = None
                    last_logs    = ""

                    for label, extra_flags in attempts_base:
                        cmd = ([PYTHON, "solve_rota.py",
                                "--input",  str(solver_input),
                                "--output", str(out_file),
                                "--seed",   str(seed),
                                "--period_name", period_name]
                               + extra_flags
                               + (["--hard_cardiac"] if hard_cardiac else []))
                        st.write(f"Running {variant_name}: **{label}**")
                        proc = subprocess.run(cmd, capture_output=True, text=True)
                        if proc.stdout: st.code(proc.stdout[-1500:])
                        if proc.stderr: st.code(proc.stderr[-1500:])
                        if proc.returncode == 0 and out_file.exists() and out_file.stat().st_size > 0:
                            succeeded  = True
                            used_label = label
                            break
                        last_logs = (proc.stderr or proc.stdout or "")[-3000:]

                    if not succeeded:
                        st.warning(f"{variant_name} failed. Logs:")
                        if last_logs: st.code(last_logs)
                        continue
                    results_local.append((variant_name, out_file.read_bytes(), used_label))

            if not results_local:
                st.error("No variants produced. Check solver logs above.")
                st.stop()

            st.session_state["draft_results"] = results_local
            st.session_state["draft_done"]    = True

        st.success(f"Produced {len(st.session_state['draft_results'])} variant(s). Download below.")

    # ── download buttons (always shown when results exist) ────────────────
    results = st.session_state.get("draft_results", [])
    if results:
        st.markdown("**Download drafted rota variants**")
        cols = st.columns(min(3, len(results)))
        for idx, (variant_name, data_bytes, used_label) in enumerate(results):
            safe_fn = (f"Rota_{variant_name.replace(' ','_').replace('(','').replace(')','')}.xlsx")
            with cols[idx % len(cols)]:
                st.download_button(
                    label=f"⬇ {variant_name}" + (f" — {used_label}" if used_label else ""),
                    data=data_bytes,
                    file_name=safe_fn,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_{idx}_{safe_fn}",
                )
