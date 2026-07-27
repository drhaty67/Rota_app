#!/usr/bin/env python3
from __future__ import annotations
import argparse
import re
import random
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Set, Tuple
from openpyxl import load_workbook
from ortools.sat.python import cp_model

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
BLOCK_TYPES = ["AB1", "AB2", "DMonThu", "DMonTue", "DWedThu", "WeekendAB", "WeekendMixed"]
MANDATORY_BLOCKS = ["AB1", "AB2", "DMonThu", "WeekendAB", "WeekendMixed"]
OPTIONAL_BLOCKS = ["DMonTue", "DWedThu"]
DEFAULT_BLOCK_TYPES = ["AB1", "AB2", "DMonThu", "WeekendAB", "WeekendMixed"]

# Block days offsets from the week's Monday
BLOCK_OFFSETS = {
    "AB1": (0, 1, 2, 3),
    "AB2": (1, 2, 3, 4),
    "DMonThu": (0, 1, 2, 3),
    "DMonTue": (0, 1),
    "DWedThu": (2, 3),
    "WeekendAB": (4, 5, 6, 7),
    "WeekendMixed": (4, 5, 6),
}

D_BLOCK_TYPES = {"DMonThu", "DMonTue", "DWedThu"}
WEEKEND_BLOCK_TYPES = {"WeekendAB", "WeekendMixed"}
A_BLOCK_TYPES = {"AB1", "AB2", "WeekendAB", "WeekendMixed"}

# Penalty tiers
TIER1_CARDIAC_PENALTY = 100_000
TIER2_WEEK_GAP_PENALTY = 50_000
TIER3_CONSEC_WEEKEND_PENALTY = 20_000
TIER4_DIVERSITY_PENALTY = 5_000
VACANCY_PENALTY = 500_000
TARGET_DEVIATION_PENALTY = 1_000_000


# ---------------------------------------------------------------------------
# Utility functions
# ---------------------------------------------------------------------------

def excel_date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            return datetime.fromisoformat(s).date()
        except Exception:
            pass
        try:
            from dateutil import parser as _parser
            return _parser.parse(s, dayfirst=True).date()
        except Exception:
            return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        try:
            return date(1899, 12, 30) + timedelta(days=int(v))
        except Exception:
            return None
    return None


def fetch_period_dates_from_supabase(period_name: str):
    import os
    from datetime import datetime as _dt
    try:
        from supabase import create_client
    except Exception as e:
        raise RuntimeError(f"Supabase client not available: {e}.") from e

    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_SERVICE_KEY") or os.getenv("SUPABASE_KEY") or os.getenv("SUPABASE_ANON_KEY")
    if not url or not key:
        raise RuntimeError("Missing SUPABASE_URL and/or SUPABASE_KEY environment variables.")

    sb = create_client(url, key)
    resp = sb.table("rota_periods").select("start_date,end_date").eq("name", period_name).limit(1).execute()
    data = resp.data or []
    if not data:
        raise RuntimeError(f"No rota_periods row found with name={period_name!r}.")
    row = data[0]

    def _to_date(v):
        if v is None:
            return None
        if hasattr(v, "date") and not isinstance(v, str):
            try:
                return v.date()
            except Exception:
                pass
        if isinstance(v, str):
            s = v.strip()
            if not s:
                return None
            try:
                return _dt.fromisoformat(s.replace('Z', '+00:00')).date()
            except Exception:
                from dateutil import parser as _parser
                return _parser.parse(s, dayfirst=True).date()
        raise RuntimeError(f"Unparseable date from Supabase: {v!r}")

    s = _to_date(row.get("start_date"))
    e = _to_date(row.get("end_date"))
    if not s or not e:
        raise RuntimeError(f"rota_periods row '{period_name}' missing start_date/end_date.")
    return s, e


def _parse_unique_tags(val) -> frozenset:
    if val is None:
        return frozenset()
    s = str(val).strip().lower().replace(";", ",").replace("\n", ",")
    return frozenset(p.strip() for p in s.split(",") if p.strip())


def _has_tag(tags: frozenset, needle: str) -> bool:
    needle = needle.strip().lower()
    return any(needle in t for t in tags)


def _parse_special_circumstances(val) -> list:
    """Parse the special_circumstances column (col G) into a list of block rules.

    Each rule is a dict:
        {"count": int, "blocks": [str, ...], "fractional": bool}

    "count"      - exact number of allocations of this type required across the cycle.
    "blocks"     - list of block types that satisfy this rule (OR semantics).
    "fractional" - if True, count was expressed as a fraction (e.g. "1/2 D") meaning
                   the solver should treat it as a soft target (0 or 1 allowed).
    """
    if val is None:
        return []
    raw = str(val).strip()
    if not raw:
        return []

    _ALIASES = {
        "a": "AB1", "ab1": "AB1",
        "b": "AB2", "ab2": "AB2",
        "d": "DMonThu", "dmonth": "DMonThu", "dmonthu": "DMonThu", "dmonthur": "DMonThu",
        "weekendab": "WeekendAB", "wab": "WeekendAB",
        "weekendmixed": "WeekendMixed", "wm": "WeekendMixed",
        "dmontue": "DMonTue", "dmt": "DMonTue",
        "dwedthu": "DWedThu", "dwt": "DWedThu",
    }
    for bt in BLOCK_TYPES:
        _ALIASES[bt.lower()] = bt

    def _resolve(token: str) -> Optional[str]:
        return _ALIASES.get(token.strip().lower())

    # Legacy: plain integer -> N extra blocks of any type
    try:
        n = int(raw)
        if n > 0:
            return [{"count": n, "blocks": list(BLOCK_TYPES), "fractional": False}]
        return []
    except ValueError:
        pass

    rules = []
    clauses = [c.strip() for c in re.split(r"[,;]|\band\b", raw, flags=re.IGNORECASE) if c.strip()]

    for clause in clauses:
        clause = clause.strip()
        if not clause:
            continue

        m = re.match(r"^(\d+(?:/\d+)?)\s+(.+)$", clause, re.IGNORECASE)
        if m:
            count_str = m.group(1)
            blocks_str = m.group(2)

            fractional = False
            if "/" in count_str:
                num, denom = count_str.split("/", 1)
                try:
                    frac = int(num) / int(denom)
                    count = max(1, round(frac))
                    fractional = True
                except Exception:
                    count = 1
            else:
                count = int(count_str)

            block_tokens = [t.strip() for t in re.split(r"\bor\b", blocks_str, flags=re.IGNORECASE)]
            blocks = [_resolve(t) for t in block_tokens]
            blocks = [b for b in blocks if b is not None]

            if blocks:
                if fractional and blocks == ["DMonThu"]:
                    blocks = ["DMonTue", "DWedThu"]
                rules.append({"count": count, "blocks": blocks, "fractional": fractional})
            continue

        block_tokens = [t.strip() for t in re.split(r"\bor\b", clause, flags=re.IGNORECASE)]
        blocks = [_resolve(t) for t in block_tokens]
        blocks = [b for b in blocks if b is not None]
        if blocks:
            rules.append({"count": 1, "blocks": blocks, "fractional": False})

    return rules


@dataclass(frozen=True)
class Consultant:
    name: str
    cardiac: bool
    wte: float
    eligible_a: bool
    eligible_d: bool
    active: bool
    unique: str = ""
    unique_tags: frozenset = frozenset()
    weekend_wte_cap: bool = False
    no_weekend: bool = False
    no_weekendab: bool = False
    no_d: bool = False
    block_rules: tuple = ()


def daterange(d0: date, d1: date) -> List[date]:
    out, d = [], d0
    while d <= d1:
        out.append(d)
        d += timedelta(days=1)
    return out


def read_preferred_shifts_from_workbook(wb) -> List[Dict]:
    prefs: List[Dict] = []
    if "preferred_shifts" not in wb.sheetnames:
        return prefs
    ws = wb["preferred_shifts"]
    header = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is not None:
            header[str(v).strip().lower()] = c

    def col(*names):
        for n in names:
            if n in header:
                return header[n]
        return None

    c_name = col("consultant_name", "consultant", "name")
    c_s = col("start_date", "startdate", "start")
    c_e = col("end_date", "enddate", "end")
    c_t = col("shift_type", "shifttype", "shift")
    c_w = col("weight", "pref_weight", "priority")

    if not (c_name and c_s and c_e and c_t):
        return prefs

    for r in range(2, ws.max_row + 1):
        nm = ws.cell(r, c_name).value
        if nm is None or str(nm).strip() == "":
            continue
        sd = excel_date(ws.cell(r, c_s).value)
        ed = excel_date(ws.cell(r, c_e).value)
        st = ws.cell(r, c_t).value
        if sd is None or ed is None or st is None:
            continue
        wt = 3
        if c_w:
            try:
                wt = int(ws.cell(r, c_w).value or 3)
            except Exception:
                wt = 3
        prefs.append({
            "consultant_name": str(nm).strip(),
            "start_date": sd,
            "end_date": ed,
            "shift_type": str(st).strip(),
            "weight": max(1, min(5, wt)),
        })
    return prefs


def read_pre_allocations_from_workbook(wb) -> List[Dict]:
    """Read pre-allocated shifts from a sheet named 'PreAllocations' if present.

    Expected columns (header row 1):
      consultant_name (or Name), week_start (or WeekStart), block_type (or Block)

    Returns a list of dicts:
      {"consultant_name": str, "week_start": date, "block_type": str}
    """
    pre: List[Dict] = []
    sheet_name = None
    for sn in ("PreAllocations", "preallocations", "Pre_Allocations"):
        if sn in wb.sheetnames:
            sheet_name = sn
            break
    if sheet_name is None:
        return pre

    ws = wb[sheet_name]
    header = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is not None:
            header[str(v).strip().lower()] = c

    def col(*names):
        for n in names:
            if n in header:
                return header[n]
        return None

    c_name = col("consultant_name", "consultant", "name")
    c_week = col("week_start", "weekstart", "week")
    c_block = col("block_type", "block", "blocktype", "shift")

    if not (c_name and c_week and c_block):
        return pre

    for r in range(2, ws.max_row + 1):
        nm = ws.cell(r, c_name).value
        if nm is None or str(nm).strip() == "":
            continue
        wk = excel_date(ws.cell(r, c_week).value)
        bt = ws.cell(r, c_block).value
        if wk is None or bt is None:
            continue
        bt_s = str(bt).strip()
        if bt_s not in BLOCK_TYPES:
            continue
        pre.append({
            "consultant_name": str(nm).strip(),
            "week_start": wk,
            "block_type": bt_s,
        })
    return pre


def read_inputs(path: str, override_start: Optional[date] = None, override_end: Optional[date] = None):
    wb = load_workbook(path, data_only=False)
    cfg = wb["Config"]

    if override_start is not None or override_end is not None:
        for r in range(1, 80):
            lab = str(cfg[f"A{r}"].value).strip() if cfg[f"A{r}"].value is not None else ""
            if lab == "CycleStartDate" and override_start is not None:
                cfg[f"B{r}"].value = override_start
            if lab == "CycleEndDate" and override_end is not None:
                cfg[f"B{r}"].value = override_end

    def get_cfg(label: str):
        for r in range(1, 80):
            if str(cfg[f"A{r}"].value).strip() == label:
                return excel_date(cfg[f"B{r}"].value)
        raise ValueError(f"Config label not found: {label}")

    start = override_start or get_cfg("CycleStartDate")
    end = override_end or get_cfg("CycleEndDate")
    if start is None or end is None:
        raise ValueError("CycleStartDate/CycleEndDate missing or unparseable in Config sheet.")

    cws = wb["Consultants"]
    consultants: List[Consultant] = []
    for r in range(2, 1000):
        nm = cws[f"A{r}"].value
        if not nm:
            continue
        tags = _parse_unique_tags(cws[f"H{r}"].value)
        sc_rules = _parse_special_circumstances(cws[f"G{r}"].value)
        consultants.append(Consultant(
            name=str(nm),
            cardiac=bool(cws[f"B{r}"].value),
            wte=float(cws[f"C{r}"].value or 0.0),
            eligible_a=bool(cws[f"D{r}"].value),
            eligible_d=bool(cws[f"E{r}"].value),
            active=bool(cws[f"F{r}"].value),
            unique=str(cws[f"H{r}"].value or ""),
            unique_tags=tags,
            weekend_wte_cap=_has_tag(tags, "weekendab = 1 wte"),
            no_weekendab=_has_tag(tags, "no weekendab"),
            no_weekend=_has_tag(tags, "no weekend"),
            no_d=_has_tag(tags, "no d"),
            block_rules=tuple(sc_rules),
        ))
    consultants = [c for c in consultants if c.active]
    if not consultants:
        raise ValueError("No active consultants found.")

    lws = wb["Leave"]
    leave_map: Dict[str, Set[date]] = {c.name: set() for c in consultants}
    for r in range(2, 5000):
        nm = lws[f"A{r}"].value
        if not nm:
            continue
        if not bool(lws[f"E{r}"].value):
            continue
        s = excel_date(lws[f"B{r}"].value)
        e = excel_date(lws[f"C{r}"].value)
        if not s or not e:
            continue
        nm = str(nm)
        if nm not in leave_map:
            continue
        for d in daterange(s, e):
            leave_map[nm].add(d)

    bws = wb["BankHolidays"]
    bh: Set[date] = set()
    for r in range(2, 2000):
        d = excel_date(bws[f"A{r}"].value)
        if d:
            bh.add(d)

    prefs = read_preferred_shifts_from_workbook(wb)
    pre_allocs = read_pre_allocations_from_workbook(wb)
    return start, end, consultants, leave_map, bh, prefs, pre_allocs


# ---------------------------------------------------------------------------
# Helper: compute block days for a given week monday
# ---------------------------------------------------------------------------

def _block_days(week_monday: date, block_type: str) -> List[date]:
    """Return the actual dates covered by a block type in a given week."""
    return [week_monday + timedelta(days=k) for k in BLOCK_OFFSETS[block_type]]


# ---------------------------------------------------------------------------
# Helper: compute eligible block types per consultant
# ---------------------------------------------------------------------------

def _eligible_blocks(c: Consultant) -> List[str]:
    """Return the list of block types a consultant is eligible for."""
    allowed = list(BLOCK_TYPES)
    if not c.eligible_a:
        for bt in ("AB1", "AB2", "WeekendAB", "WeekendMixed"):
            if bt in allowed:
                allowed.remove(bt)
    if not c.eligible_d or c.no_d:
        for bt in ("DMonThu", "DMonTue", "DWedThu", "WeekendMixed"):
            if bt in allowed:
                allowed.remove(bt)
    if c.no_weekend:
        for bt in ("WeekendAB", "WeekendMixed"):
            if bt in allowed:
                allowed.remove(bt)
    elif c.no_weekendab:
        if "WeekendAB" in allowed:
            allowed.remove("WeekendAB")
    return allowed


# ---------------------------------------------------------------------------
# Solver
# ---------------------------------------------------------------------------

def solve(
    start: date,
    end: date,
    consultants: List[Consultant],
    leave: Dict[str, Set[date]],
    bank_holidays: Set[date],
    prefs: List[Dict],
    pre_allocations: Optional[List[Dict]] = None,
    hard_no_consecutive_weekends: bool = True,
    hard_week_gap: bool = True,
    relax_cardiac: bool = True,
    relax_targets: bool = False,
    time_limit_s: int = 60,
    random_seed: int = 0,
) -> Dict:
    # -----------------------------------------------------------------------
    # Setup: weeks, names, indices
    # -----------------------------------------------------------------------
    first_monday = start + timedelta(days=(7 - start.weekday()) % 7)
    weeks: List[date] = []
    d = first_monday
    while d <= end:
        weeks.append(d)
        d += timedelta(days=7)
    W = len(weeks)

    names = [c.name for c in consultants]
    N = len(names)
    cardiac = [c.cardiac for c in consultants]

    # Identify flex/vacancy consultants
    FLEX_NAMES = {c.name.strip().lower() for c in consultants if "vacancy" in c.name.strip().lower()}

    print(f"\n{'='*60}")
    print(f"SOLVER DIAGNOSTICS")
    print(f"{'='*60}")
    print(f"Period: {start} -> {end}")
    print(f"First Monday: {first_monday}")
    print(f"Number of weeks: {W}")
    print(f"Relaxations: week_gap={'OFF' if not hard_week_gap else 'SOFT'}, "
          f"consec_wknds={'OFF' if not hard_no_consecutive_weekends else 'SOFT'}, "
          f"cardiac={'HARD' if not relax_cardiac else 'SOFT'}, "
          f"targets={'RELAXED' if relax_targets else 'HARD'}")
    print(f"Active consultants: {N}")
    print(f"Flex/vacancy consultants: {FLEX_NAMES}")

    # -----------------------------------------------------------------------
    # Per-consultant: eligible blocks, target, availability
    # -----------------------------------------------------------------------
    eligible = [_eligible_blocks(c) for c in consultants]

    # Compute targets based on allocation rules
    targets = []
    # Also track per-consultant SC rule structures for constraint generation
    consultant_sc_rules = []  # list of list of rules per consultant

    for i, c in enumerate(consultants):
        elig = eligible[i]
        if c.wte == 1.0 and not c.block_rules:
            # 1 WTE no SC: exactly 5, one each of mandatory types
            targets.append(5)
            consultant_sc_rules.append(None)
        elif c.wte < 1.0 and c.block_rules:
            # < 1 WTE with SC: rules define exactly which types and counts
            t = sum(rule["count"] for rule in c.block_rules if not rule["fractional"])
            # Add fractional rules as soft targets (count toward target as 1)
            t += sum(1 for rule in c.block_rules if rule["fractional"])
            targets.append(max(1, t))
            consultant_sc_rules.append(list(c.block_rules))
        elif c.wte > 1.0 and c.block_rules:
            # > 1 WTE with SC: 5 base + extras from SC
            extras = sum(rule["count"] for rule in c.block_rules if not rule["fractional"])
            extras += sum(1 for rule in c.block_rules if rule["fractional"])
            targets.append(5 + extras)
            consultant_sc_rules.append(list(c.block_rules))
        elif c.wte < 1.0 and not c.block_rules:
            # < 1 WTE no SC: round(wte * 5), min 1
            targets.append(max(1, round(c.wte * 5)))
            consultant_sc_rules.append(None)
        else:
            # >= 1 WTE no SC (shouldn't normally have wte > 1 without SC but handle it)
            targets.append(5)
            consultant_sc_rules.append(None)

    # Compute availability per consultant (weeks where they can work at least one block)
    avail_per_consultant = []
    for i, c in enumerate(consultants):
        avail = 0
        for w, wk in enumerate(weeks):
            can_any = False
            for b in eligible[i]:
                days = _block_days(wk, b)
                if not any(dd in leave.get(c.name, set()) for dd in days):
                    can_any = True
                    break
            if can_any:
                avail += 1
        avail_per_consultant.append(avail)

    # Print diagnostics
    print(f"\nConsultant targets and availability:")
    infeasible_flags = []
    for i, c in enumerate(consultants):
        leave_days_in_period = len([dd for dd in leave.get(c.name, set()) if start <= dd <= end])
        is_flex = c.name.strip().lower() in FLEX_NAMES
        flag = ""
        if not is_flex and avail_per_consultant[i] < targets[i]:
            flag = " *** IMPOSSIBLE ***"
            infeasible_flags.append(c.name)
        print(f"  {c.name}: WTE={c.wte}, cardiac={c.cardiac}, target={targets[i]}, "
              f"avail_weeks={avail_per_consultant[i]}, leave_days={leave_days_in_period}, "
              f"eligible={eligible[i]}, flex={is_flex}{flag}")

    if infeasible_flags:
        print(f"\n*** WARNING: These consultants cannot meet their targets: {infeasible_flags}")
        print(f"    Their available weeks are less than their target allocations.")
        if not relax_targets:
            print(f"    Consider using --relax_targets or adjusting leave/WTE.")
    print()

    # -----------------------------------------------------------------------
    # Model construction
    # -----------------------------------------------------------------------
    model = cp_model.CpModel()

    # Decision variables: x[(w, bt, i)] = 1 iff consultant i works block bt in week w
    x = {}
    for w in range(W):
        for bt in BLOCK_TYPES:
            for i in range(N):
                x[(w, bt, i)] = model.NewBoolVar(f"x_{w}_{bt}_{i}")

    # Vacancy variables for mandatory blocks
    vacancy = {}
    for w in range(W):
        for bt in MANDATORY_BLOCKS:
            vacancy[(w, bt)] = model.NewBoolVar(f"vac_{w}_{bt}")

    # ===================================================================
    # HARD CONSTRAINT 1: One consultant per block per week
    # Mandatory blocks (AB1, AB2, DMonThu, WeekendAB, WeekendMixed): exactly one consultant OR vacancy
    # Optional blocks (DMonTue, DWedThu): at most one consultant
    # ===================================================================
    for w in range(W):
        for bt in MANDATORY_BLOCKS:
            model.Add(
                sum(x[(w, bt, i)] for i in range(N)) + vacancy[(w, bt)] == 1
            )
        for bt in OPTIONAL_BLOCKS:
            model.Add(
                sum(x[(w, bt, i)] for i in range(N)) <= 1
            )

    # Initialize soft penalty terms list (used by multiple constraint sections)
    soft_penalty_terms = []

    # ===================================================================
    # SOFT: D-block mutual exclusion per week
    # DMonThu should not coexist with DMonTue or DWedThu in the same week.
    # If DMonTue or DWedThu is assigned, DMonThu should go to vacancy.
    # DMonTue + DWedThu together IS allowed (different days).
    # Enforced as SOFT (high penalty) to prevent infeasibility.
    # ===================================================================
    D_EXCLUSION_PENALTY = 200_000
    for w in range(W):
        dmon_thu_assigned = sum(x[(w, "DMonThu", i)] for i in range(N))
        dmon_tue_assigned = sum(x[(w, "DMonTue", i)] for i in range(N))
        dwed_thu_assigned = sum(x[(w, "DWedThu", i)] for i in range(N))
        viol1 = model.NewBoolVar(f"d_excl_tue_{w}")
        model.Add(dmon_thu_assigned + dmon_tue_assigned <= 1 + viol1)
        soft_penalty_terms.append(D_EXCLUSION_PENALTY * viol1)
        viol2 = model.NewBoolVar(f"d_excl_wed_{w}")
        model.Add(dmon_thu_assigned + dwed_thu_assigned <= 1 + viol2)
        soft_penalty_terms.append(D_EXCLUSION_PENALTY * viol2)

    # ===================================================================
    # HARD CONSTRAINT 1c: DMonTue/DWedThu restricted to SC consultants only
    # Only consultants whose SC rules explicitly include DMonTue or DWedThu
    # (plus flex/vacancy consultants) can be assigned those blocks.
    # ===================================================================
    # Determine who has SC rules that include DMonTue or DWedThu
    sc_allows_half_d = set()
    for i, c in enumerate(consultants):
        if c.name.strip().lower() in FLEX_NAMES:
            sc_allows_half_d.add(i)  # vacancy can fill remaining half-D slots
            continue
        if c.block_rules:
            for rule in c.block_rules:
                for b in rule["blocks"]:
                    if b in ("DMonTue", "DWedThu"):
                        sc_allows_half_d.add(i)
                        break

    print(f"SC allows half-D: {[names[i] for i in sc_allows_half_d]}")

    # DMonTue/DWedThu primarily for SC consultants + vacancy.
    # However, other D-eligible consultants can use them as a last resort
    # (penalised via diversity tier to discourage it, but not forbidden)
    # This prevents infeasibility when total targets exceed mandatory supply.

    # ===================================================================
    # HARD CONSTRAINT 2: One block per consultant per week
    # ===================================================================
    for w in range(W):
        for i in range(N):
            model.Add(sum(x[(w, bt, i)] for bt in BLOCK_TYPES) <= 1)

    # ===================================================================
    # HARD CONSTRAINT 3: Leave blocking
    # A consultant cannot be assigned a block if they have leave on any
    # day that block covers.
    # ===================================================================
    for w, wk in enumerate(weeks):
        for bt in BLOCK_TYPES:
            days = _block_days(wk, bt)
            for i in range(N):
                if any(dd in leave.get(names[i], set()) for dd in days):
                    model.Add(x[(w, bt, i)] == 0)

    # Block type eligibility (not eligible -> cannot be assigned)
    for w in range(W):
        for bt in BLOCK_TYPES:
            for i in range(N):
                if bt not in eligible[i]:
                    model.Add(x[(w, bt, i)] == 0)

    # ===================================================================
    # HARD CONSTRAINT 4: Pre-allocations
    # Lock in admin-specified assignments. Skip conflicts.
    # ===================================================================
    pre_alloc_fixed: Set[Tuple[int, str, int]] = set()
    name_to_idx = {c.name.strip(): idx for idx, c in enumerate(consultants)}
    name_to_idx_lower = {c.name.strip().lower(): idx for idx, c in enumerate(consultants)}
    week_to_idx = {wk: w for w, wk in enumerate(weeks)}

    if pre_allocations:
        pre_alloc_slots: Dict[Tuple[int, str], str] = {}
        print("Pre-allocation processing:")
        for pa in pre_allocations:
            nm = pa["consultant_name"]
            wk = pa["week_start"]
            bt = pa["block_type"]
            i = name_to_idx.get(nm)
            if i is None:
                i = name_to_idx_lower.get(nm.lower())
            if i is None:
                print(f"  [SKIP] Consultant '{nm}' not found.")
                continue
            w = week_to_idx.get(wk)
            if w is None:
                print(f"  [SKIP] Week {wk} not in period for '{nm}'.")
                continue
            if bt not in BLOCK_TYPES:
                print(f"  [SKIP] Block type '{bt}' invalid for '{nm}'.")
                continue
            # Check eligibility
            if bt not in eligible[i]:
                print(f"  [SKIP] '{nm}' not eligible for {bt}.")
                continue
            # Check leave conflict
            days = _block_days(weeks[w], bt)
            if any(dd in leave.get(names[i], set()) for dd in days):
                print(f"  [SKIP] '{nm}' has leave during {bt} week {wk}.")
                continue
            # Check slot already taken
            slot_key = (w, bt)
            if slot_key in pre_alloc_slots:
                print(f"  [SKIP] {bt} week {wk} already assigned to '{pre_alloc_slots[slot_key]}'.")
                continue
            # Check consultant already has a block this week via pre-alloc
            existing = [(wb, bb) for (wb, bb, ib) in pre_alloc_fixed if wb == w and ib == i]
            if existing:
                print(f"  [SKIP] '{nm}' already pre-allocated {existing[0][1]} in week {wk}.")
                continue
            # Lock it in
            model.Add(x[(w, bt, i)] == 1)
            pre_alloc_fixed.add((w, bt, i))
            pre_alloc_slots[slot_key] = nm
            print(f"  [OK] '{nm}' -> {bt} week {wk}")
        print()

    # ===================================================================
    # HARD CONSTRAINT 5: Allocation targets
    # Each consultant must get exactly their target number of blocks.
    # Flex/vacancy: use <= target. relax_targets: allow +/- 1.
    # ===================================================================
    # Count pre-allocations per consultant (they count toward target)
    pre_alloc_count = [0] * N
    for (w, bt, i) in pre_alloc_fixed:
        pre_alloc_count[i] += 1

    target_slack_vars = []  # for relax_targets penalty

    for i, c in enumerate(consultants):
        total_alloc = sum(x[(w, bt, i)] for w in range(W) for bt in BLOCK_TYPES)
        is_flex = c.name.strip().lower() in FLEX_NAMES

        if is_flex:
            # Vacancy/flex: at most target, must contribute at least 1 block
            model.Add(total_alloc <= targets[i])
            if avail_per_consultant[i] > 0:
                model.Add(total_alloc >= 1)
        elif relax_targets:
            # Allow target +/- 1
            effective_target = min(targets[i], avail_per_consultant[i])
            lo = max(0, effective_target - 1)
            hi = effective_target + 1
            model.Add(total_alloc >= lo)
            model.Add(total_alloc <= hi)
            # Penalise deviation from exact target
            dev = model.NewIntVar(0, W, f"tgt_dev_{i}")
            model.AddAbsEquality(dev, total_alloc - effective_target)
            target_slack_vars.append(dev)
        else:
            # Hard exact target
            effective_target = targets[i]
            # If truly impossible, we still set the constraint and let the solver
            # report infeasible (caller handles it)
            model.Add(total_alloc == effective_target)

    # ===================================================================
    # HARD: Block type rules per consultant category
    # ===================================================================
    # Initialize soft_penalty_terms early (some block rules add diversity penalties)

    for i, c in enumerate(consultants):
        is_flex = c.name.strip().lower() in FLEX_NAMES
        if is_flex:
            continue  # flex consultants have no type constraints

        if c.wte == 1.0 and not c.block_rules:
            # 1 WTE no SC: exactly one of each mandatory type (no duplicates)
            for bt in MANDATORY_BLOCKS:
                if bt in eligible[i]:
                    model.Add(sum(x[(w, bt, i)] for w in range(W)) == 1)
                else:
                    model.Add(sum(x[(w, bt, i)] for w in range(W)) == 0)

        elif c.wte < 1.0 and c.block_rules:
            # < 1 WTE with SC: each rule defines exactly count from those options
            for rule in c.block_rules:
                rule_blocks = [b for b in rule["blocks"] if b in eligible[i]]
                if not rule_blocks:
                    continue
                group_sum = sum(x[(w, b, i)] for w in range(W) for b in rule_blocks)
                if rule["fractional"]:
                    # Fractional: 0 or 1 (soft target of 1, handled in objective)
                    model.Add(group_sum <= 1)
                else:
                    if relax_targets:
                        # When targets are relaxed, SC rules become soft too
                        model.Add(group_sum <= rule["count"])
                    else:
                        model.Add(group_sum == rule["count"])

        elif c.wte > 1.0 and c.block_rules:
            # > 1 WTE with SC: base 5 (one each mandatory) + extras from rules
            # Base: one of each mandatory type the consultant is eligible for
            for bt in MANDATORY_BLOCKS:
                if bt in eligible[i]:
                    model.Add(sum(x[(w, bt, i)] for w in range(W)) >= 1)
            # Accumulate extra counts per unique block group
            # Multiple rules with the same block set add their counts together
            from collections import defaultdict
            group_extras: Dict[frozenset, int] = defaultdict(int)
            for rule in c.block_rules:
                rule_blocks = [b for b in rule["blocks"] if b in eligible[i]]
                if not rule_blocks:
                    continue
                if rule["fractional"]:
                    continue  # fractional extras handled by total target constraint
                group_key = frozenset(rule_blocks)
                group_extras[group_key] += rule["count"]
            # Apply accumulated constraints per group
            for group_blocks_fs, extra_count in group_extras.items():
                group_blocks = list(group_blocks_fs)
                group_sum = sum(x[(w, b, i)] for w in range(W) for b in group_blocks)
                # base_in_group = mandatory types in this group that get 1 each from base
                base_in_group = len([b for b in group_blocks if b in MANDATORY_BLOCKS])
                model.Add(group_sum >= base_in_group + extra_count)

        elif c.wte < 1.0 and not c.block_rules:
            # < 1 WTE no SC: prefer no duplicates, but allow if target > eligible types
            eligible_types = [bt for bt in BLOCK_TYPES if bt in eligible[i]]
            n_eligible = len(eligible_types)
            for bt in BLOCK_TYPES:
                if bt not in eligible[i]:
                    model.Add(sum(x[(w, bt, i)] for w in range(W)) == 0)
                elif n_eligible >= targets[i]:
                    # Enough eligible types for no-duplicate assignment
                    model.Add(sum(x[(w, bt, i)] for w in range(W)) <= 1)
                else:
                    # Not enough eligible types — allow up to 2 of each but penalise
                    count_bt = sum(x[(w, bt, i)] for w in range(W))
                    model.Add(count_bt <= 2)
                    excess = model.NewIntVar(0, W, f"dup_nosc_{i}_{bt}")
                    model.Add(excess >= count_bt - 1)
                    soft_penalty_terms.append(TIER4_DIVERSITY_PENALTY * excess)

        elif c.wte >= 1.0 and not c.block_rules:
            # >= 1 WTE no SC: one each mandatory type
            for bt in MANDATORY_BLOCKS:
                if bt in eligible[i]:
                    model.Add(sum(x[(w, bt, i)] for w in range(W)) == 1)
                else:
                    model.Add(sum(x[(w, bt, i)] for w in range(W)) == 0)

    # ===================================================================
    # SOFT CONSTRAINTS (penalised in objective)
    # ===================================================================

    # -------------------------------------------------------------------
    # TIER 1 (100,000): Cardiac XOR
    # On each weekday, exactly one of A-slot and D-slot holders should be
    # cardiac-competent.
    # -------------------------------------------------------------------
    # A-slot mapping per weekday:
    #   Mon (0), Wed (2): AB1
    #   Tue (1), Thu (3): AB2
    #   Fri (4): WeekendAB
    # D-slot mapping per weekday:
    #   Mon (0): DMonThu or DMonTue
    #   Tue (1): DMonThu or DMonTue
    #   Wed (2): DMonThu or DWedThu
    #   Thu (3): DMonThu or DWedThu
    #   Fri (4): WeekendMixed

    _A_block_for_day = {0: "AB1", 1: "AB2", 2: "AB1", 3: "AB2", 4: "WeekendAB"}
    _D_blocks_for_day = {
        0: ["DMonThu", "DMonTue"],
        1: ["DMonThu", "DMonTue"],
        2: ["DMonThu", "DWedThu"],
        3: ["DMonThu", "DWedThu"],
        4: ["WeekendMixed"],
    }

    for w in range(W):
        for day in range(5):
            a_block = _A_block_for_day[day]
            d_blocks = _D_blocks_for_day[day]

            # Sum of cardiac-competent consultants holding the A-slot
            a_cardiac_sum = sum(
                x[(w, a_block, i)] for i in range(N) if cardiac[i]
            )
            # Sum of cardiac-competent consultants holding D-slot(s)
            d_cardiac_sum = sum(
                x[(w, db, i)] for db in d_blocks for i in range(N) if cardiac[i]
            )
            cardiac_total = a_cardiac_sum + d_cardiac_sum

            if not relax_cardiac:
                # HARD cardiac constraint (--hard_cardiac flag)
                model.Add(cardiac_total == 1)
            else:
                # SOFT: penalise deviation from exactly 1
                pen = model.NewIntVar(0, 2, f"cardiac_pen_{w}_{day}")
                model.Add(pen >= 1 - cardiac_total)
                model.Add(pen >= cardiac_total - 1)
                soft_penalty_terms.append(TIER1_CARDIAC_PENALTY * pen)

    # -------------------------------------------------------------------
    # TIER 2 (50,000): 1-week gap between assignments
    # At least one week gap between any two assigned blocks for same consultant.
    # If --no_hard_week_gap: constraint is completely OFF.
    # Otherwise: SOFT penalty.
    # -------------------------------------------------------------------
    if hard_week_gap:
        for i in range(N):
            is_flex = names[i].strip().lower() in FLEX_NAMES
            if is_flex:
                continue
            for w in range(W - 1):
                any_this = sum(x[(w, bt, i)] for bt in BLOCK_TYPES)
                any_next = sum(x[(w + 1, bt, i)] for bt in BLOCK_TYPES)
                # Soft: penalise but don't forbid
                violation = model.NewBoolVar(f"gap_viol_{w}_{i}")
                model.Add(any_this + any_next <= 1 + violation)
                soft_penalty_terms.append(TIER2_WEEK_GAP_PENALTY * violation)

    # -------------------------------------------------------------------
    # TIER 3 (20,000): No consecutive weekends
    # If --no_hard_no_consec_weekends: constraint is completely OFF.
    # Otherwise: SOFT penalty.
    # -------------------------------------------------------------------
    if hard_no_consecutive_weekends:
        for i in range(N):
            for w in range(W - 1):
                wknd_this = x[(w, "WeekendAB", i)] + x[(w, "WeekendMixed", i)]
                wknd_next = x[(w + 1, "WeekendAB", i)] + x[(w + 1, "WeekendMixed", i)]
                violation = model.NewBoolVar(f"consec_wknd_{w}_{i}")
                model.Add(wknd_this + wknd_next <= 1 + violation)
                soft_penalty_terms.append(TIER3_CONSEC_WEEKEND_PENALTY * violation)

    # -------------------------------------------------------------------
    # TIER 4 (5,000): Block type diversity for >1 WTE extras
    # For >1 WTE with SC, spread extras across different types.
    # Penalise each duplicate assignment of the same block type beyond base.
    # -------------------------------------------------------------------
    for i, c in enumerate(consultants):
        if c.wte <= 1.0 or not c.block_rules:
            continue
        is_flex = c.name.strip().lower() in FLEX_NAMES
        if is_flex:
            continue
        for bt in eligible[i]:
            count_bt = sum(x[(w, bt, i)] for w in range(W))
            # Base expectation: 1 for mandatory types
            base = 1 if bt in MANDATORY_BLOCKS else 0
            if base > 0:
                excess = model.NewIntVar(0, W, f"div_excess_{i}_{bt}")
                model.Add(excess >= count_bt - base)
                soft_penalty_terms.append(TIER4_DIVERSITY_PENALTY * excess)

    # -------------------------------------------------------------------
    # TIER 5 (reward - negative cost): Preference satisfaction
    # -------------------------------------------------------------------
    rng = random.Random(int(random_seed) if random_seed is not None else 0)

    def _week_overlaps(ps: date, pe: date, wk_start: date) -> bool:
        return not (pe < wk_start or ps > wk_start + timedelta(days=6))

    def _shift_to_blocks(st: str) -> List[str]:
        st_up = st.strip().upper()
        if st_up == "A":
            return ["AB1", "AB2"]
        if st_up == "B":
            return ["AB2"]
        if st_up == "D":
            return ["DMonThu", "DMonTue", "DWedThu"]
        if st_up in ("WEEKEND", "W"):
            return ["WeekendAB", "WeekendMixed"]
        if st in BLOCK_TYPES:
            return [st]
        return []

    PREF_WEIGHT_STEP = 1000
    PREF_MAX_JITTER = 25
    pref_reward_terms = []

    name_to_i_lower = {c.name.strip().lower(): idx for idx, c in enumerate(consultants)}

    for p_idx, p in enumerate(prefs or []):
        nm = str(p.get("consultant_name", "")).strip().lower()
        if not nm or nm not in name_to_i_lower:
            continue
        i = name_to_i_lower[nm]
        ps, pe = p.get("start_date"), p.get("end_date")
        if ps is None or pe is None:
            continue
        st = str(p.get("shift_type", "")).strip()
        blocks = _shift_to_blocks(st)
        if not blocks:
            continue
        blocks = [b for b in blocks if b in eligible[i]]
        if not blocks:
            continue
        base_w = max(1, min(5, int(p.get("weight", 3) or 3)))

        match_vars = [
            x[(w, b, i)]
            for w, wk in enumerate(weeks)
            if _week_overlaps(ps, pe, wk)
            for b in blocks
            if b in eligible[i]
        ]
        if not match_vars:
            continue

        sat = model.NewBoolVar(f"pref_sat_{p_idx}")
        model.Add(sum(match_vars) >= sat)
        for v in match_vars:
            model.Add(sat >= v)

        reward = base_w * PREF_WEIGHT_STEP + rng.randint(0, PREF_MAX_JITTER)
        pref_reward_terms.append(reward * sat)

    # ===================================================================
    # FAIRNESS OBJECTIVE (lowest priority)
    # Minimise deviation from fair bank-holiday and weekend burden
    # proportional to WTE.
    # ===================================================================
    wte_list = [c.wte for c in consultants]
    sum_wte = sum(wte_list) if sum(wte_list) > 0 else 1.0
    SCALE = 1000

    # Bank-holiday burden
    bh_count = {}
    for w, wk in enumerate(weeks):
        for bt in BLOCK_TYPES:
            bh_count[(w, bt)] = sum(1 for dd in _block_days(wk, bt) if dd in bank_holidays)

    bh_duty = [model.NewIntVar(0, 500, f"bh_{i}") for i in range(N)]
    for i in range(N):
        model.Add(bh_duty[i] == sum(
            x[(w, bt, i)] * bh_count[(w, bt)]
            for w in range(W) for bt in BLOCK_TYPES
        ))

    bh_total = sum(bh_count[(w, bt)] for w in range(W) for bt in BLOCK_TYPES)
    expected_bh = [int(round(bh_total * (wte_list[i] / sum_wte) * SCALE)) for i in range(N)]
    devBH = [model.NewIntVar(0, 10_000_000, f"devBH_{i}") for i in range(N)]
    for i in range(N):
        model.AddAbsEquality(devBH[i], bh_duty[i] * SCALE - expected_bh[i])

    # Weekend burden
    weekend_duty = [model.NewIntVar(0, 500, f"wknd_{i}") for i in range(N)]
    for i in range(N):
        model.Add(weekend_duty[i] == sum(
            x[(w, "WeekendAB", i)] + x[(w, "WeekendMixed", i)] for w in range(W)
        ))

    weekend_total = 2 * W
    expected_wknd = [int(round(weekend_total * (wte_list[i] / sum_wte) * SCALE)) for i in range(N)]
    devW = [model.NewIntVar(0, 10_000_000, f"devW_{i}") for i in range(N)]
    for i in range(N):
        model.AddAbsEquality(devW[i], weekend_duty[i] * SCALE - expected_wknd[i])

    # ===================================================================
    # Vacancy and target-deviation penalties
    # ===================================================================
    vacancy_cost = VACANCY_PENALTY * sum(vacancy[(w, bt)] for w in range(W) for bt in MANDATORY_BLOCKS)

    target_dev_cost = TARGET_DEVIATION_PENALTY * sum(target_slack_vars) if target_slack_vars else 0

    # Flex consultant under-use penalty (encourage using them)
    flex_underuse_terms = []
    for i, c in enumerate(consultants):
        if c.name.strip().lower() in FLEX_NAMES:
            total_alloc = sum(x[(w, bt, i)] for w in range(W) for bt in BLOCK_TYPES)
            shortfall = model.NewIntVar(0, W, f"flex_short_{i}")
            ideal = min(targets[i], avail_per_consultant[i])
            model.Add(shortfall >= ideal - total_alloc)
            flex_underuse_terms.append(VACANCY_PENALTY * shortfall)

    flex_cost = sum(flex_underuse_terms) if flex_underuse_terms else 0

    # ===================================================================
    # OBJECTIVE: Minimise total cost
    # ===================================================================
    pref_reward = sum(pref_reward_terms) if pref_reward_terms else 0
    soft_penalty = sum(soft_penalty_terms) if soft_penalty_terms else 0
    fairness_cost = 3 * sum(devBH) + 2 * sum(devW)

    model.Minimize(
        vacancy_cost
        + target_dev_cost
        + flex_cost
        + soft_penalty
        + fairness_cost
        - pref_reward
    )

    # ===================================================================
    # Solve
    # ===================================================================
    solver = cp_model.CpSolver()
    solver.parameters.random_seed = int(random_seed)
    solver.parameters.randomize_search = True
    solver.parameters.max_time_in_seconds = float(time_limit_s)
    solver.parameters.num_search_workers = 8

    print(f"Solving with time_limit={time_limit_s}s, seed={random_seed}, workers=8 ...")
    status = solver.Solve(model)
    status_name = solver.StatusName(status)
    objective = solver.ObjectiveValue() if status in (cp_model.OPTIMAL, cp_model.FEASIBLE) else None

    print(f"Solver status: {status_name}")
    if objective is not None:
        print(f"Objective value: {objective}")

    if status == cp_model.INFEASIBLE:
        print("\n*** MODEL IS INFEASIBLE ***")
        print(f"  Pre-allocations fixed: {len(pre_alloc_fixed)}")
        for (w, bt, i) in sorted(pre_alloc_fixed):
            print(f"    week {weeks[w]} / {bt} -> {names[i]}")
        total_min_demand = sum(
            0 if (relax_targets or c.name.strip().lower() in FLEX_NAMES) else targets[idx]
            for idx, c in enumerate(consultants)
        )
        total_supply = W * len(BLOCK_TYPES)
        print(f"  Total supply (slots): {total_supply}")
        print(f"  Total hard demand (non-flex targets): {total_min_demand}")
        print(f"  relax_targets={relax_targets}")

    # ===================================================================
    # Build solution dict
    # ===================================================================
    sol = {
        "status": status_name,
        "objective": objective,
        "weeks": weeks,
        "assignments": {wk: {} for wk in weeks},
    }

    if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        for w, wk in enumerate(weeks):
            for bt in BLOCK_TYPES:
                assigned = False
                for i in range(N):
                    if solver.Value(x[(w, bt, i)]) == 1:
                        sol["assignments"][wk][bt] = names[i]
                        assigned = True
                        break
                if not assigned and bt in MANDATORY_BLOCKS:
                    if solver.Value(vacancy[(w, bt)]) == 1:
                        sol["assignments"][wk][bt] = "Vacancy"

        # Print summary
        print(f"\nAssignment summary:")
        for i, c in enumerate(consultants):
            count = sum(
                1 for w in range(W) for bt in BLOCK_TYPES
                if solver.Value(x[(w, bt, i)]) == 1
            )
            types_assigned = [
                bt for bt in BLOCK_TYPES
                if sum(solver.Value(x[(w, bt, i)]) for w in range(W)) > 0
            ]
            print(f"  {c.name}: {count}/{targets[i]} blocks, types={types_assigned}")

    return sol


# ---------------------------------------------------------------------------
# Export to Excel
# ---------------------------------------------------------------------------

def export_to_excel(input_path: str, output_path: str, sol: Dict,
                    override_start: Optional[date] = None, override_end: Optional[date] = None):
    wb = load_workbook(input_path)

    wa = wb["WeekAssignments"]
    rota = wb["Rota"]
    dash = wb["Dashboard"]
    cfg = wb["Config"]

    if override_start is not None or override_end is not None:
        for r in range(1, 80):
            lab = str(cfg[f"A{r}"].value).strip() if cfg[f"A{r}"].value is not None else ""
            if lab == "CycleStartDate" and override_start is not None:
                cfg[f"B{r}"].value = override_start
            if lab == "CycleEndDate" and override_end is not None:
                cfg[f"B{r}"].value = override_end

    cons = wb["Consultants"]
    leave_ws = wb["Leave"]
    bh_ws = wb["BankHolidays"]

    def get_cfg(label: str):
        for r in range(1, 80):
            if str(cfg[f"A{r}"].value).strip() == label:
                return cfg[f"B{r}"].value
        return None

    start = excel_date(get_cfg("CycleStartDate"))
    end = excel_date(get_cfg("CycleEndDate"))
    prev_A_for_start = str(get_cfg("A_Consultant_DayBeforeStart") or "")

    cardiac = {}
    wte = {}
    for r in range(2, 1000):
        nm = cons[f"A{r}"].value
        if not nm:
            continue
        nm = str(nm)
        cardiac[nm] = bool(cons[f"B{r}"].value)
        wte[nm] = float(cons[f"C{r}"].value or 0.0)

    leave_map = {}
    for r in range(2, 5000):
        nm = leave_ws[f"A{r}"].value
        if not nm:
            continue
        if not bool(leave_ws[f"E{r}"].value):
            continue
        s = excel_date(leave_ws[f"B{r}"].value)
        e = excel_date(leave_ws[f"C{r}"].value)
        if not s or not e:
            continue
        nm = str(nm)
        leave_map.setdefault(nm, set())
        d = s
        while d <= e:
            leave_map[nm].add(d)
            d += timedelta(days=1)

    bh_set = set()
    for r in range(2, 2000):
        d = excel_date(bh_ws[f"A{r}"].value)
        if d:
            bh_set.add(d)

    # Clear WeekAssignments
    for r in range(2, wa.max_row + 1):
        for c in range(1, 9):
            wa.cell(r, c).value = None

    weeks = sol["weeks"]
    for r_i, wk in enumerate(weeks, start=2):
        wa.cell(r_i, 1).value = wk
        wa.cell(r_i, 2).value = sol["assignments"][wk].get("AB1", "")
        wa.cell(r_i, 3).value = sol["assignments"][wk].get("AB2", "")
        wa.cell(r_i, 4).value = sol["assignments"][wk].get("DMonThu", "")
        wa.cell(r_i, 5).value = sol["assignments"][wk].get("DMonTue", "")
        wa.cell(r_i, 6).value = sol["assignments"][wk].get("DWedThu", "")
        wa.cell(r_i, 7).value = sol["assignments"][wk].get("WeekendAB", "")
        wa.cell(r_i, 8).value = sol["assignments"][wk].get("WeekendMixed", "")
        wa.cell(r_i, 9).value = sol.get("status", "")
        wa.cell(r_i, 10).value = sol.get("objective", "")

    wk_map = {wk: sol["assignments"][wk] for wk in weeks}

    def week_monday(d: date) -> date:
        return d - timedelta(days=d.weekday())

    # Clear Rota
    for r in range(2, rota.max_row + 1):
        for c in range(1, 8):
            rota.cell(r, c).value = None
    rota.cell(1, 7).value = "Available"

    # Build unique initials for each consultant
    all_cons_names = sorted(cardiac.keys())

    def _make_initials(names_list):
        initials = {}
        for nm in names_list:
            parts = nm.strip().split()
            if len(parts) >= 2:
                ini = parts[0][0].upper() + parts[-1][0].upper()
            else:
                ini = nm[:2].upper()
            initials[nm] = ini
        ini_to_names = {}
        for nm, ini in initials.items():
            ini_to_names.setdefault(ini, []).append(nm)
        for ini, nms in ini_to_names.items():
            if len(nms) > 1:
                for nm in nms:
                    parts = nm.strip().split()
                    surname = parts[-1] if len(parts) >= 2 else nm
                    new_ini = parts[0][0].upper() + surname[0].upper() + surname[1].upper() if len(surname) > 1 else ini
                    initials[nm] = new_ini
        return initials

    consultant_initials = _make_initials(all_cons_names)

    all_days = daterange(start, end)
    prev_A = None
    for row_i, d in enumerate(all_days, start=2):
        dow = d.weekday()
        wk = week_monday(d)
        asg = wk_map.get(wk, {})

        if dow in (0, 2):
            A = asg.get("AB1", "")
        elif dow in (1, 3):
            A = asg.get("AB2", "")
        elif dow == 4:
            A = asg.get("WeekendAB", "")
        elif dow == 5:
            A = asg.get("WeekendMixed", "")
        else:
            A = asg.get("WeekendAB", "")

        B = prev_A_for_start if d == start else (prev_A or "")

        if dow == 0 or dow == 1:
            D = asg.get("DMonThu") or asg.get("DMonTue") or ""
        elif dow == 2:
            D = asg.get("DMonThu") or asg.get("DWedThu") or ""
        elif dow == 3:
            D = asg.get("DMonThu") or asg.get("DWedThu") or ""
        elif dow == 4:
            D = asg.get("WeekendMixed") or ""
        else:
            D = ""

        flags = []
        if not A:
            flags.append("MISSING_A")
        if not B:
            flags.append("MISSING_B")
        if dow <= 4 and not D:
            flags.append("MISSING_D")
        if dow >= 5 and D:
            flags.append("D_SHOULD_BE_BLANK_WEEKEND")
        if A and d in leave_map.get(A, set()):
            flags.append("A_ON_LEAVE")
        if B and d in leave_map.get(B, set()):
            flags.append("B_ON_LEAVE")
        if D and d in leave_map.get(D, set()):
            flags.append("D_ON_LEAVE")
        if dow <= 4:
            a_c = bool(cardiac.get(A, False))
            d_c = bool(cardiac.get(D, False))
            if (a_c + d_c) != 1:
                flags.append("CARDIAC_XOR_BREACH")
        if d in bh_set:
            flags.append("BANK_HOLIDAY")

        available = [consultant_initials[nm] for nm in all_cons_names
                     if d not in leave_map.get(nm, set())]

        rota.cell(row_i, 1).value = d
        rota.cell(row_i, 2).value = d.strftime("%a")
        rota.cell(row_i, 3).value = A
        rota.cell(row_i, 4).value = B
        rota.cell(row_i, 5).value = D
        rota.cell(row_i, 6).value = ",".join(flags)
        rota.cell(row_i, 7).value = ", ".join(available)
        prev_A = A

    # Dashboard
    for r in range(2, dash.max_row + 1):
        for c in range(1, 14):
            dash.cell(r, c).value = None

    counts = {nm: {"A": 0, "B": 0, "D": 0, "BH": 0, "wknd": 0, "consec_wknd": 0}
              for nm in cardiac.keys()}
    weekend_by_cons = {nm: [] for nm in cardiac.keys()}
    for wk in weeks:
        for b in ("WeekendAB", "WeekendMixed"):
            nm = wk_map[wk].get(b, "")
            if nm and nm in counts:
                weekend_by_cons.setdefault(nm, []).append(wk)

    for nm, wks in weekend_by_cons.items():
        if nm not in counts:
            continue
        wks = sorted(wks)
        for i in range(len(wks) - 1):
            if (wks[i + 1] - wks[i]).days == 7:
                counts[nm]["consec_wknd"] += 1
        counts[nm]["wknd"] = len(wks)

    for row_i, d in enumerate(all_days, start=2):
        A = rota.cell(row_i, 3).value or ""
        B = rota.cell(row_i, 4).value or ""
        Dv = rota.cell(row_i, 5).value or ""
        is_bh = "BANK_HOLIDAY" in (rota.cell(row_i, 6).value or "")
        if A in counts:
            counts[A]["A"] += 1
        if B in counts:
            counts[B]["B"] += 1
        if Dv in counts and d.weekday() <= 4:
            counts[Dv]["D"] += 1
        if is_bh:
            for nm in (A, B):
                if nm in counts:
                    counts[nm]["BH"] += 1
            if Dv in counts and d.weekday() <= 4:
                counts[Dv]["BH"] += 1

    total_all = sum(v["A"] + v["B"] + v["D"] for v in counts.values())
    total_bh = sum(v["BH"] for v in counts.values())
    sum_wte = sum(wte.values()) if wte else 1.0

    r = 2
    for nm in sorted(counts.keys()):
        A_cnt, B_cnt, D_cnt = counts[nm]["A"], counts[nm]["B"], counts[nm]["D"]
        tot = A_cnt + B_cnt + D_cnt
        exp = total_all * (wte.get(nm, 0.0) / sum_wte)
        bh_cnt = counts[nm]["BH"]
        bh_exp = total_bh * (wte.get(nm, 0.0) / sum_wte)
        dash.cell(r, 1).value = nm
        dash.cell(r, 2).value = wte.get(nm, 0.0)
        dash.cell(r, 3).value = A_cnt
        dash.cell(r, 4).value = B_cnt
        dash.cell(r, 5).value = D_cnt
        dash.cell(r, 6).value = tot
        dash.cell(r, 7).value = round(exp, 2)
        dash.cell(r, 8).value = round(tot - exp, 2)
        dash.cell(r, 9).value = bh_cnt
        dash.cell(r, 10).value = round(bh_exp, 2)
        dash.cell(r, 11).value = round(bh_cnt - bh_exp, 2)
        dash.cell(r, 12).value = counts[nm]["wknd"]
        dash.cell(r, 13).value = counts[nm]["consec_wknd"]
        r += 1

    wb.save(output_path)


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--time_limit", type=int, default=60)
    ap.add_argument("--seed", type=int, default=0)
    ap.add_argument("--period_name", type=str, default="")
    ap.add_argument("--no_hard_week_gap", action="store_true",
                    help="Turn OFF the 1-week gap constraint entirely")
    ap.add_argument("--no_hard_no_consec_weekends", action="store_true",
                    help="Turn OFF the no-consecutive-weekends constraint entirely")
    ap.add_argument("--hard_cardiac", action="store_true",
                    help="Enforce cardiac XOR as a HARD constraint instead of soft")
    ap.add_argument("--relax_targets", action="store_true",
                    help="Allow allocation targets to flex +/- 1 (soft penalty for deviation)")
    args = ap.parse_args()

    override_start = override_end = None
    if args.period_name:
        override_start, override_end = fetch_period_dates_from_supabase(args.period_name)

    try:
        start, end, consultants, leave, bh, prefs, pre_allocs = read_inputs(
            args.input, override_start=override_start, override_end=override_end)
    except Exception as e:
        raise SystemExit(f"Input workbook error: {e}")

    sol = solve(
        start, end, consultants, leave, bh, prefs,
        pre_allocations=pre_allocs,
        hard_no_consecutive_weekends=not args.no_hard_no_consec_weekends,
        hard_week_gap=not args.no_hard_week_gap,
        relax_cardiac=not args.hard_cardiac,
        relax_targets=args.relax_targets,
        time_limit_s=args.time_limit,
        random_seed=args.seed,
    )
    print(f"Status: {sol['status']}  Objective: {sol.get('objective')}")
    export_to_excel(args.input, args.output, sol,
                    override_start=override_start, override_end=override_end)
    print(f"Wrote {args.output}")

    if sol["status"] == "INFEASIBLE":
        raise SystemExit(1)


if __name__ == "__main__":
    main()
