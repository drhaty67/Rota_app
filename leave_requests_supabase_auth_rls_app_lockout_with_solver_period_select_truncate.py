import streamlit as st
from datetime import date, datetime
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
from supabase import create_client, Client
import subprocess
import tempfile
from pathlib import Path

st.set_page_config(page_title="Rota Leave Requests + Solver", layout="wide")

# -----------------------------
# Secrets
# -----------------------------
SUPABASE_URL = st.secrets.get("SUPABASE_URL", "")
SUPABASE_ANON_KEY = st.secrets.get("SUPABASE_ANON_KEY", "")  # public key
ALLOWED_EMAIL_DOMAIN = st.secrets.get("ALLOWED_EMAIL_DOMAIN", "")  # optional

if not SUPABASE_URL or not SUPABASE_ANON_KEY:
    st.error("Missing SUPABASE_URL or SUPABASE_ANON_KEY in Streamlit secrets.")
    st.stop()

ALLOWED_TYPES = ["Annual", "Study", "NOC"]

# -----------------------------
# Supabase client helpers
# -----------------------------
def base_client() -> Client:
    return create_client(SUPABASE_URL, SUPABASE_ANON_KEY)

def authed_client(access_token: str) -> Client:
    c = base_client()
    c.postgrest.auth(access_token)  # enforce RLS
    return c

def validate_dates(s: date, e: date) -> str | None:
    if e < s:
        return "Date to cannot be earlier than Date from."
    return None

def overlap(a_start: date, a_end: date, b_start: date, b_end: date) -> bool:
    return a_start <= b_end and b_start <= a_end

# -----------------------------
# Auth UI
# -----------------------------
st.title("Rota Leave Requests")
st.write(
    "Supabase Auth + Row Level Security (RLS), with lockout after a rota period is published. "
    "Admins can approve leave, compile leave into the Excel template, and run the rota solver to draft the rota."
)

if "sb_session" not in st.session_state:
    st.session_state["sb_session"] = None

with st.sidebar:
    st.header("Sign in")
    email = st.text_input("Email", value="").strip().lower()
    password = st.text_input("Password", type="password")

    if ALLOWED_EMAIL_DOMAIN and email and not email.endswith("@" + ALLOWED_EMAIL_DOMAIN):
        st.warning(f"Email must end with @{ALLOWED_EMAIL_DOMAIN}.")

    c = base_client()

    col1, col2 = st.columns(2)
    with col1:
        if st.button("Sign in", use_container_width=True):
            if not email or not password:
                st.error("Enter email and password.")
            elif ALLOWED_EMAIL_DOMAIN and not email.endswith("@" + ALLOWED_EMAIL_DOMAIN):
                st.error("Email domain not permitted.")
            else:
                try:
                    res = c.auth.sign_in_with_password({"email": email, "password": password})
                    st.session_state["sb_session"] = res.session.model_dump() if res.session else None
                    st.rerun()
                except Exception:
                    st.error("Sign-in failed. Check credentials and confirmation status.")

    with col2:
        if st.button("Sign up", use_container_width=True):
            if not email or not password:
                st.error("Enter email and password.")
            elif ALLOWED_EMAIL_DOMAIN and not email.endswith("@" + ALLOWED_EMAIL_DOMAIN):
                st.error("Email domain not permitted.")
            else:
                try:
                    _ = c.auth.sign_up({"email": email, "password": password})
                    st.success("Sign-up created. If email confirmation is enabled, confirm via email then sign in.")
                except Exception:
                    st.error("Sign-up failed. Email may already exist or password may be too weak.")

    if st.session_state["sb_session"]:
        if st.button("Sign out", use_container_width=True):
            st.session_state["sb_session"] = None
            st.rerun()

sess = st.session_state["sb_session"]
if not sess:
    st.info("Please sign in to continue.")
    st.stop()

access_token = sess["access_token"]
user_id = sess["user"]["id"]
user_email = sess["user"]["email"]

db = authed_client(access_token)

# -----------------------------
# Admin detection
# -----------------------------
def is_rota_admin() -> bool:
    try:
        r = db.table("rota_admins").select("user_id").limit(1).execute()
        return bool(r.data)
    except Exception:
        return False

is_admin = is_rota_admin()

with st.sidebar:
    st.markdown("---")
    st.write(f"Signed in as: {user_email}")
    st.write("Role: Rota admin" if is_admin else "Role: Consultant")

# -----------------------------
# Published periods
# -----------------------------
def fetch_periods() -> pd.DataFrame:
    resp = db.table("rota_periods").select("*").order("start_date").execute()
    data = resp.data or []
    dfp = pd.DataFrame(data)
    if not dfp.empty:
        dfp["start_date"] = pd.to_datetime(dfp["start_date"]).dt.date
        dfp["end_date"] = pd.to_datetime(dfp["end_date"]).dt.date
        dfp["published_at"] = pd.to_datetime(dfp["published_at"], errors="coerce")
        dfp["created_at"] = pd.to_datetime(dfp["created_at"], errors="coerce")
    return dfp

periods = fetch_periods()

def is_locked_for_range(s: date, e: date) -> bool:
    if periods.empty:
        return False
    pubs = periods[periods["is_published"] == True]
    if pubs.empty:
        return False
    for _, p in pubs.iterrows():
        if overlap(s, e, p["start_date"], p["end_date"]):
            return True
    return False

st.subheader("0) Published rota periods (lockout)")
if periods.empty:
    st.info("No rota periods configured yet.")
else:
    st.dataframe(periods[["id","name","start_date","end_date","is_published","published_at"]],
                 use_container_width=True, hide_index=True)
    if not periods[periods["is_published"] == True].empty:
        st.caption("Requests overlapping published periods cannot be created/edited/deleted by consultants.")

# -----------------------------
# Data access
# -----------------------------
def fetch_leave_requests() -> pd.DataFrame:
    resp = db.table("leave_requests").select("*").order("start_date").execute()
    data = resp.data or []
    df = pd.DataFrame(data)
    if not df.empty:
        df["start_date"] = pd.to_datetime(df["start_date"]).dt.date
        df["end_date"] = pd.to_datetime(df["end_date"]).dt.date
        df["created_at"] = pd.to_datetime(df["created_at"])
        df["updated_at"] = pd.to_datetime(df["updated_at"])
    return df

df = fetch_leave_requests()

# -----------------------------
# Add request
# -----------------------------
st.subheader("1) Submit a leave request")
with st.form("add_form"):
    c1, c2, c3 = st.columns([2, 1, 1])
    with c1:
        consultant_name = st.text_input("Consultant name (as it should appear on the rota)", value="")
        leave_type = st.selectbox("Leave type", options=ALLOWED_TYPES)
        notes = st.text_input("Notes (optional)", value="")
    with c2:
        start_date = st.date_input("Date from", value=date.today())
    with c3:
        end_date = st.date_input("Date to", value=date.today())
    submitted = st.form_submit_button("Submit request")

if submitted:
    err = validate_dates(start_date, end_date)
    if err:
        st.error(err)
    elif not consultant_name.strip():
        st.error("Consultant name is required.")
    else:
        if (not is_admin) and is_locked_for_range(start_date, end_date):
            st.error("This rota period has been published. You cannot submit leave that overlaps a published period.")
        else:
            try:
                db.table("leave_requests").insert({
                    "consultant_name": consultant_name.strip(),
                    "requester_id": user_id,
                    "requester_email": user_email,
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat(),
                    "leave_type": leave_type,
                    "approved": False,
                    "notes": notes.strip()
                }).execute()
                st.success("Request submitted (pending approval).")
                st.rerun()
            except Exception:
                st.error("Insert failed (likely locked by published period or RLS not applied).")

# -----------------------------
# View requests
# -----------------------------
st.subheader("2) Your requests")
df = fetch_leave_requests()
if df.empty:
    st.info("No requests found.")
else:
    view_cols = ["id", "consultant_name", "start_date", "end_date", "leave_type", "approved", "notes", "updated_at"]
    st.dataframe(df[view_cols], use_container_width=True, hide_index=True)

# -----------------------------
# Edit/delete pending
# -----------------------------
st.subheader("3) Edit or delete a pending request")
pending = df[df["approved"] == False].copy() if not df.empty else pd.DataFrame()
if pending.empty:
    st.info("No pending requests available to edit/delete.")
else:
    selected_id = st.selectbox("Select request", options=pending["id"].tolist())
    row = pending[pending["id"] == selected_id].iloc[0]

    locked_row = (not is_admin) and is_locked_for_range(row["start_date"], row["end_date"])
    if locked_row:
        st.warning("This request overlaps a published period and is locked. Only an admin can modify it.")

    with st.form("edit_form"):
        e1, e2, e3 = st.columns([2, 1, 1])
        with e1:
            new_name = st.text_input("Consultant name", value=row["consultant_name"], disabled=locked_row)
            new_type = st.selectbox("Leave type", options=ALLOWED_TYPES,
                                    index=ALLOWED_TYPES.index(row["leave_type"]), disabled=locked_row)
            new_notes = st.text_input("Notes", value=row.get("notes") or "", disabled=locked_row)
        with e2:
            new_start = st.date_input("Date from", value=row["start_date"], disabled=locked_row)
        with e3:
            new_end = st.date_input("Date to", value=row["end_date"], disabled=locked_row)

        csave, cdel = st.columns([1, 1])
        with csave:
            save = st.form_submit_button("Save changes", disabled=locked_row)
        with cdel:
            delete = st.form_submit_button("Delete request", disabled=locked_row)

    if save:
        err = validate_dates(new_start, new_end)
        if err:
            st.error(err)
        elif not new_name.strip():
            st.error("Consultant name is required.")
        else:
            if (not is_admin) and is_locked_for_range(new_start, new_end):
                st.error("This overlaps a published period and cannot be modified.")
            else:
                try:
                    db.table("leave_requests").update({
                        "consultant_name": new_name.strip(),
                        "start_date": new_start.isoformat(),
                        "end_date": new_end.isoformat(),
                        "leave_type": new_type,
                        "notes": new_notes.strip()
                    }).eq("id", selected_id).execute()
                    st.success("Updated.")
                    st.rerun()
                except Exception:
                    st.error("Update failed (likely locked/published, or RLS not installed).")

    if delete:
        try:
            db.table("leave_requests").delete().eq("id", selected_id).execute()
            st.success("Deleted.")
            st.rerun()
        except Exception:
            st.error("Delete failed (likely locked/published, or RLS not installed).")

# -----------------------------
# Admin actions
# -----------------------------
st.subheader("4) Admin actions")
if not is_admin:
    st.info("Admin actions are available to rota administrators only.")
    st.stop()

# ---- Approvals ----
st.markdown("#### Approvals")
df_admin = fetch_leave_requests()
pending_admin = df_admin[df_admin["approved"] == False].copy() if not df_admin.empty else pd.DataFrame()
if pending_admin.empty:
    st.write("No pending requests.")
else:
    st.dataframe(pending_admin[["id","consultant_name","requester_email","start_date","end_date","leave_type","notes","created_at"]],
                 use_container_width=True, hide_index=True)
    approve_id = st.selectbox("Select pending request ID", options=pending_admin["id"].tolist(), key="appr")
    colA, colB = st.columns([1, 1])
    with colA:
        if st.button("Approve selected", use_container_width=True):
            db.table("leave_requests").update({"approved": True}).eq("id", approve_id).execute()
            st.success("Approved.")
            st.rerun()
    with colB:
        if st.button("Reject (delete) selected", use_container_width=True):
            db.table("leave_requests").delete().eq("id", approve_id).execute()
            st.success("Rejected (deleted).")
            st.rerun()

# ---- Draft rota with solver ----
st.markdown("---")
st.markdown("#### Draft rota (compile approved leave → run solver → download solved workbook)")

st.write(
    "1) Select the rota period you are publishing."
    "2) Upload your base rota workbook template."
    "3) The app will write **approved leave** that falls within the selected period into the `Leave` sheet, "
    "then run the solver. Drafting is only permitted when the selected period is **published**.")

# Refresh periods for this section
periods = fetch_periods()

if periods.empty:
    st.error("No rota periods found. Create a rota period first, then publish it before drafting.")
else:
    # Admins can see all periods; sort newest first for convenience
    periods_sorted = periods.sort_values(["start_date"], ascending=False).copy()
    # Show label with publish status
    periods_sorted["label"] = periods_sorted.apply(
        lambda r: f"{r['name']} ({r['start_date']} → {r['end_date']}) — {'PUBLISHED' if r['is_published'] else 'unpublished'}",
        axis=1
    )
    sel_label = st.selectbox("Select rota period", options=periods_sorted["label"].tolist(), key="draft_period_sel")
    sel_row = periods_sorted[periods_sorted["label"] == sel_label].iloc[0]

    period_id = sel_row["id"]
    period_start = sel_row["start_date"]
    period_end = sel_row["end_date"]
    period_published = bool(sel_row["is_published"])

    st.caption(f"Selected period window: {period_start} to {period_end}.")

    if not period_published:
        st.warning("Drafting is disabled because the selected rota period is not published.")
        st.stop()

    template = st.file_uploader("Upload base rota workbook (.xlsx)", type=["xlsx"], key="solver_tmpl")
    only_approved = st.checkbox("Use approved leave only", value=True)

    solver_script = Path("solve_rota.py")  # must exist in your repo on Streamlit Cloud

    if not solver_script.exists():
        st.warning("`solve_rota.py` not found in the app repository. Add it to your GitHub repo to enable drafting.")
    else:
        relax_week_gap = st.checkbox("Allow fallback: relax 1-week gap constraint", value=True)
        relax_no_consec_weekends = st.checkbox("Allow fallback: relax no-consecutive-weekends constraint", value=True)

        if template is None:
            st.info("Upload the base rota workbook to enable drafting.")
        else:
            if st.button("Draft rota now"):
                export_df = fetch_leave_requests()

                if only_approved:
                    export_df = export_df[export_df["approved"] == True].copy()

                # --- Validate leave alignment to selected period window ---
                # We include ONLY leave fully inside the window, but we hard-fail if any approved leave overlaps the window
                # while extending outside it (partial overlap), because that usually indicates a period boundary issue.
                if not export_df.empty:
                    # Identify leave that overlaps the period window at all
                    overlaps_period = export_df.apply(
                        lambda r: overlap(r["start_date"], r["end_date"], period_start, period_end), axis=1
                    )
                    df_overlap = export_df[overlaps_period].copy()

                    # Partial-outside = overlaps but not fully contained
                    fully_contained = df_overlap.apply(
                        lambda r: (r["start_date"] >= period_start) and (r["end_date"] <= period_end), axis=1
                    )
                    df_partial = df_overlap[~fully_contained].copy()

                    # Optional: force truncation of partial-overlap leave to the selected period window
                    force_truncate = st.checkbox(
                        "Force draft by truncating leave that partially overlaps the selected period",
                        value=False,
                        help=(
                            "If enabled, approved leave that overlaps the period but extends outside it will be "
                            "clipped to the period boundaries for drafting. The original leave record is not changed."
                        ),
                        key="force_truncate_partial"
                    )

                    if not df_partial.empty and not force_truncate:
                        st.error(
                            "Cannot draft: at least one APPROVED leave request partially overlaps the selected rota period "
                            "but extends outside the period window. Enable truncation or correct the leave dates/period boundaries."
                        )
                        st.dataframe(
                            df_partial[["consultant_name","requester_email","start_date","end_date","leave_type","approved"]],
                            use_container_width=True, hide_index=True
                        )
                        st.stop()

                    if not df_partial.empty and force_truncate:
                        df_trunc = df_partial.copy()
                        df_trunc["orig_start_date"] = df_trunc["start_date"]
                        df_trunc["orig_end_date"] = df_trunc["end_date"]
                        df_trunc["start_date"] = df_trunc["start_date"].apply(lambda d: max(d, period_start))
                        df_trunc["end_date"] = df_trunc["end_date"].apply(lambda d: min(d, period_end))

                        st.warning(
                            "Proceeding with draft using truncated leave for the selected period window. "
                            "The table below shows original vs truncated dates."
                        )
                        st.dataframe(
                            df_trunc[["consultant_name","requester_email","leave_type","approved",
                                      "orig_start_date","orig_end_date","start_date","end_date"]],
                            use_container_width=True, hide_index=True
                        )

                        # Replace partial rows with truncated versions for drafting only
                        df_overlap.loc[df_trunc.index, "start_date"] = df_trunc["start_date"]
                        df_overlap.loc[df_trunc.index, "end_date"] = df_trunc["end_date"]

                        # After truncation, treat them as contained
                        fully_contained = df_overlap.apply(
                            lambda r: (r["start_date"] >= period_start) and (r["end_date"] <= period_end), axis=1
                        )

                    # For solver input, include only leave fully inside the window (post-truncation if enabled)
                    export_df = df_overlap[fully_contained].copy()

                try:
                    # Build solver input workbook (template + Leave sheet)
                    wb = load_workbook(BytesIO(template.getvalue()))
                    if "Leave" not in wb.sheetnames:
                        st.error("Workbook must contain a sheet named 'Leave'.")
                        st.stop()

                    lws = wb["Leave"]

                    # Clear existing Leave rows
                    for r in range(2, 5000):
                        if lws[f"A{r}"].value in (None, ""):
                            break
                        for col in ("A","B","C","D","E"):
                            lws[f"{col}{r}"].value = None

                    # Write leave rows
                    r = 2
                    if not export_df.empty:
                        export_df = export_df.sort_values(["start_date","consultant_name"], na_position="last")
                        for _, rec in export_df.iterrows():
                            lws[f"A{r}"].value = rec["consultant_name"]
                            lws[f"B{r}"].value = rec["start_date"]
                            lws[f"C{r}"].value = rec["end_date"]
                            lws[f"D{r}"].value = rec["leave_type"]
                            lws[f"E{r}"].value = bool(rec["approved"])
                            r += 1

                    with tempfile.TemporaryDirectory() as td:
                        td_path = Path(td)
                        solver_input = td_path / "Rota_Master_WITH_Leave.xlsx"
                        solver_output = td_path / "Rota_Solved.xlsx"
                        wb.save(solver_input)

                        attempts = [
                            ("Strict (all hard constraints)",
                             ["python", "solve_rota.py", "--input", str(solver_input), "--output", str(solver_output)])
                        ]
                        if relax_week_gap:
                            attempts.append(("Relax week-gap constraint (--no_hard_week_gap)",
                                             ["python", "solve_rota.py", "--input", str(solver_input), "--output", str(solver_output),
                                              "--no_hard_week_gap"]))
                        if relax_no_consec_weekends:
                            attempts.append(("Relax no-consecutive-weekends constraint (--no_hard_no_consec_weekends)",
                                             ["python", "solve_rota.py", "--input", str(solver_input), "--output", str(solver_output),
                                              "--no_hard_no_consec_weekends"]))
                        if relax_week_gap and relax_no_consec_weekends:
                            attempts.append(("Relax BOTH (--no_hard_week_gap + --no_hard_no_consec_weekends)",
                                             ["python", "solve_rota.py", "--input", str(solver_input), "--output", str(solver_output),
                                              "--no_hard_week_gap", "--no_hard_no_consec_weekends"]))

                        success = False
                        for label, cmd in attempts:
                            st.write(f"Running solver: **{label}**")
                            proc = subprocess.run(cmd, capture_output=True, text=True)

                            if proc.stdout and proc.stdout.strip():
                                st.code(proc.stdout[-4000:])
                            if proc.stderr and proc.stderr.strip():
                                st.code(proc.stderr[-4000:])

                            if proc.returncode == 0 and solver_output.exists() and solver_output.stat().st_size > 0:
                                st.success(f"Solver succeeded: {label}")
                                success = True
                                break
                            else:
                                st.warning(f"Solver did not succeed: {label}")

                        if not success:
                            st.error(
                                "All solver attempts failed. Review logs above. "
                                "Consider revising leave requests or relaxing additional constraints in the solver."
                            )
                            st.stop()

                        st.download_button(
                            "Download drafted rota (Rota_Solved.xlsx)",
                            data=solver_output.read_bytes(),
                            file_name="Rota_Solved.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                except Exception as e:
                    st.exception(e)
