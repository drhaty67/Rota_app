import streamlit as st
from datetime import date
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
from supabase import create_client, Client

st.set_page_config(page_title="Rota Leave Requests", layout="wide")

# -----------------------------
# Secrets
# -----------------------------
SUPABASE_URL = st.secrets.get("SUPABASE_URL", "")
SUPABASE_ANON_KEY = st.secrets.get("SUPABASE_ANON_KEY", "")  # public key
ALLOWED_EMAIL_DOMAIN = st.secrets.get("ALLOWED_EMAIL_DOMAIN", "")  # optional, e.g. "uhl-tr.nhs.uk"

if not SUPABASE_URL or not SUPABASE_ANON_KEY:
    st.error("Missing SUPABASE_URL or SUPABASE_ANON_KEY in Streamlit secrets.")
    st.stop()

ALLOWED_TYPES = ["Annual", "Study", "NOC"]

# -----------------------------
# Supabase client helpers
# -----------------------------
def base_client() -> Client:
    return create_client(SUPABASE_URL, SUPABASE_ANON_KEY)

def authed_client(access_token: str) -> Client:
    c = base_client()
    # Ensure all PostgREST calls use the user's JWT so RLS is enforced
    c.postgrest.auth(access_token)
    return c

def validate_dates(s: date, e: date) -> str | None:
    if e < s:
        return "Date to cannot be earlier than Date from."
    return None

# -----------------------------
# Auth UI (email + password)
# -----------------------------
st.title("Rota Leave Requests")
st.write(
    "Secure leave-request capture using Supabase Auth + Row Level Security (RLS). "
    "Consultants can only see/edit their own requests. Rota admins can approve and compile to Excel."
)

if "sb_session" not in st.session_state:
    st.session_state["sb_session"] = None

with st.sidebar:
    st.header("Sign in")
    email = st.text_input("Email", value="").strip().lower()
    password = st.text_input("Password", type="password")

    if ALLOWED_EMAIL_DOMAIN and email and not email.endswith("@" + ALLOWED_EMAIL_DOMAIN):
        st.warning(f"Email must end with @{ALLOWED_EMAIL_DOMAIN}.")

    c = base_client()

    col1, col2 = st.columns(2)
    with col1:
        if st.button("Sign in", use_container_width=True):
            if not email or not password:
                st.error("Enter email and password.")
            elif ALLOWED_EMAIL_DOMAIN and not email.endswith("@" + ALLOWED_EMAIL_DOMAIN):
                st.error("Email domain not permitted.")
            else:
                try:
                    res = c.auth.sign_in_with_password({"email": email, "password": password})
                    st.session_state["sb_session"] = res.session.model_dump() if res.session else None
                    st.rerun()
                except Exception:
                    st.error("Sign-in failed. Check credentials and confirmation status.")

    with col2:
        if st.button("Sign up", use_container_width=True):
            if not email or not password:
                st.error("Enter email and password.")
            elif ALLOWED_EMAIL_DOMAIN and not email.endswith("@" + ALLOWED_EMAIL_DOMAIN):
                st.error("Email domain not permitted.")
            else:
                try:
                    _ = c.auth.sign_up({"email": email, "password": password})
                    st.success("Sign-up created. If email confirmation is enabled, confirm via email then sign in.")
                except Exception:
                    st.error("Sign-up failed. Email may already exist or password may be too weak.")

    if st.session_state["sb_session"]:
        if st.button("Sign out", use_container_width=True):
            st.session_state["sb_session"] = None
            st.rerun()

sess = st.session_state["sb_session"]
if not sess:
    st.info("Please sign in to continue.")
    st.stop()

access_token = sess["access_token"]
user_id = sess["user"]["id"]
user_email = sess["user"]["email"]

db = authed_client(access_token)

# -----------------------------
# Admin detection
# -----------------------------
def is_rota_admin() -> bool:
    # With RLS on rota_admins, only admins can read any row (policy).
    try:
        r = db.table("rota_admins").select("user_id").limit(1).execute()
        return bool(r.data)
    except Exception:
        return False

is_admin = is_rota_admin()

with st.sidebar:
    st.markdown("---")
    st.write(f"Signed in as: {user_email}")
    st.write("Role: Rota admin" if is_admin else "Role: Consultant")

# -----------------------------
# Data access
# -----------------------------
def fetch_leave_requests() -> pd.DataFrame:
    resp = db.table("leave_requests").select("*").order("start_date").execute()
    data = resp.data or []
    df = pd.DataFrame(data)
    if not df.empty:
        df["start_date"] = pd.to_datetime(df["start_date"]).dt.date
        df["end_date"] = pd.to_datetime(df["end_date"]).dt.date
        df["created_at"] = pd.to_datetime(df["created_at"])
        df["updated_at"] = pd.to_datetime(df["updated_at"])
    return df

df = fetch_leave_requests()

# -----------------------------
# Add request
# -----------------------------
st.subheader("1) Submit a leave request")
with st.form("add_form"):
    c1, c2, c3 = st.columns([2, 1, 1])
    with c1:
        consultant_name = st.text_input("Consultant name (as it should appear on the rota)", value="")
        leave_type = st.selectbox("Leave type", options=ALLOWED_TYPES)
        notes = st.text_input("Notes (optional)", value="")
    with c2:
        start_date = st.date_input("Date from", value=date.today())
    with c3:
        end_date = st.date_input("Date to", value=date.today())
    submitted = st.form_submit_button("Submit request")

if submitted:
    err = validate_dates(start_date, end_date)
    if err:
        st.error(err)
    elif not consultant_name.strip():
        st.error("Consultant name is required.")
    else:
        # RLS enforces: requester_id must equal auth.uid() and approved must be false (admin-only approval)
        try:
            db.table("leave_requests").insert({
                "consultant_name": consultant_name.strip(),
                "requester_id": user_id,
                "requester_email": user_email,
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
                "leave_type": leave_type,
                "approved": False,
                "notes": notes.strip()
            }).execute()
            st.success("Request submitted (pending approval).")
            st.rerun()
        except Exception:
            st.error("Insert failed. Confirm RLS policies were applied.")

# -----------------------------
# View requests (RLS ensures consultants only see their own)
# -----------------------------
st.subheader("2) Your requests")
if df.empty:
    st.info("No requests found.")
else:
    view_cols = ["id", "consultant_name", "start_date", "end_date", "leave_type", "approved", "notes", "updated_at"]
    st.dataframe(df[view_cols], use_container_width=True, hide_index=True)

# -----------------------------
# Edit/delete (consultants can only edit while approved = false; enforced by RLS)
# -----------------------------
st.subheader("3) Edit or delete a pending request")

pending = df[df["approved"] == False].copy() if not df.empty else pd.DataFrame()
if pending.empty:
    st.info("No pending requests available to edit/delete.")
else:
    selected_id = st.selectbox("Select request", options=pending["id"].tolist())
    row = pending[pending["id"] == selected_id].iloc[0]

    with st.form("edit_form"):
        e1, e2, e3 = st.columns([2, 1, 1])
        with e1:
            new_name = st.text_input("Consultant name", value=row["consultant_name"])
            new_type = st.selectbox("Leave type", options=ALLOWED_TYPES, index=ALLOWED_TYPES.index(row["leave_type"]))
            new_notes = st.text_input("Notes", value=row.get("notes") or "")
        with e2:
            new_start = st.date_input("Date from", value=row["start_date"])
        with e3:
            new_end = st.date_input("Date to", value=row["end_date"])

        csave, cdel = st.columns([1, 1])
        with csave:
            save = st.form_submit_button("Save changes")
        with cdel:
            delete = st.form_submit_button("Delete request")

    if save:
        err = validate_dates(new_start, new_end)
        if err:
            st.error(err)
        elif not new_name.strip():
            st.error("Consultant name is required.")
        else:
            try:
                # RLS blocks changes if approved has become true since page load
                db.table("leave_requests").update({
                    "consultant_name": new_name.strip(),
                    "start_date": new_start.isoformat(),
                    "end_date": new_end.isoformat(),
                    "leave_type": new_type,
                    "notes": new_notes.strip()
                }).eq("id", selected_id).execute()
                st.success("Updated.")
                st.rerun()
            except Exception:
                st.error("Update failed (likely already approved, or RLS not installed).")

    if delete:
        try:
            db.table("leave_requests").delete().eq("id", selected_id).execute()
            st.success("Deleted.")
            st.rerun()
        except Exception:
            st.error("Delete failed (likely already approved, or RLS not installed).")

# -----------------------------
# Admin: approve/reject + compile to Excel
# -----------------------------
st.subheader("4) Admin actions")

if not is_admin:
    st.info("Admin actions are available to rota administrators only.")
    st.stop()

st.markdown("#### Approvals")
df_admin = fetch_leave_requests()
pending_admin = df_admin[df_admin["approved"] == False].copy() if not df_admin.empty else pd.DataFrame()

if pending_admin.empty:
    st.write("No pending requests.")
else:
    st.dataframe(
        pending_admin[["id","consultant_name","requester_email","start_date","end_date","leave_type","notes","created_at"]],
        use_container_width=True,
        hide_index=True
    )

    approve_id = st.selectbox("Select pending request ID", options=pending_admin["id"].tolist())
    colA, colB = st.columns([1,1])
    with colA:
        if st.button("Approve selected", use_container_width=True):
            try:
                db.table("leave_requests").update({"approved": True}).eq("id", approve_id).execute()
                st.success("Approved.")
                st.rerun()
            except Exception:
                st.error("Approve failed. Confirm admin RLS policy is installed.")
    with colB:
        if st.button("Reject (delete) selected", use_container_width=True):
            try:
                db.table("leave_requests").delete().eq("id", approve_id).execute()
                st.success("Rejected (deleted).")
                st.rerun()
            except Exception:
                st.error("Reject failed. Confirm admin RLS policy is installed.")

st.markdown("---")
st.markdown("#### Compile approved leave into Excel workbook (download)")
st.write(
    "Upload your rota workbook template (must contain a sheet named `Leave`). "
    "The app will write **approved** leave rows only, then provide a downloadable workbook."
)

template = st.file_uploader("Upload rota workbook (.xlsx)", type=["xlsx"])

only_approved = st.checkbox("Export approved only", value=True)
if template is None:
    st.info("Upload a workbook to enable compilation.")
else:
    if st.button("Compile and generate workbook"):
        export_df = fetch_leave_requests()
        if only_approved:
            export_df = export_df[export_df["approved"] == True].copy()

        wb = load_workbook(BytesIO(template.getvalue()))
        if "Leave" not in wb.sheetnames:
            st.error("Workbook must contain a sheet named 'Leave'.")
            st.stop()
        lws = wb["Leave"]

        # Clear existing Leave rows (row 2 onward)
        for r in range(2, 5000):
            if lws[f"A{r}"].value in (None, ""):
                break
            for col in ("A","B","C","D","E"):
                lws[f"{col}{r}"].value = None

        # Write rows
        r = 2
        if not export_df.empty:
            export_df = export_df.sort_values(["start_date","consultant_name"], na_position="last")
            for _, rec in export_df.iterrows():
                lws[f"A{r}"].value = rec["consultant_name"]
                lws[f"B{r}"].value = rec["start_date"]
                lws[f"C{r}"].value = rec["end_date"]
                lws[f"D{r}"].value = rec["leave_type"]
                lws[f"E{r}"].value = bool(rec["approved"])
                r += 1

        out = BytesIO()
        wb.save(out)
        out.seek(0)

        st.download_button(
            "Download compiled workbook",
            data=out,
            file_name="Rota_Master_WITH_Leave.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
