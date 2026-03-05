#!/usr/bin/env python3
from __future__ import annotations
import argparse
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Set, Tuple
from openpyxl import load_workbook
from ortools.sat.python import cp_model

def excel_date(v) -> Optional[date]:
    """Coerce common Excel / openpyxl cell values to a python date.

    Accepts:
      - datetime/date
      - Excel serial numbers (int/float)
      - strings (YYYY-MM-DD, DD/MM/YYYY, etc.)
    Returns None if blank/unparseable.
    """
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            return datetime.fromisoformat(s).date()
        except Exception:
            pass
        try:
            from dateutil import parser as _parser
            return _parser.parse(s, dayfirst=True).date()
        except Exception:
            return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        try:
            return date(1899, 12, 30) + timedelta(days=int(v))
        except Exception:
            return None
    return None


def fetch_period_dates_from_supabase(period_name: str):
    """Fetch (start_date, end_date) from Supabase public.rota_periods by name.

    Expects env vars:
      - SUPABASE_URL
      - SUPABASE_KEY   (anon key or service role key). For server-side admin runs, service role is OK.
    Returns (start_date: date, end_date: date).
    """
    import os
    from datetime import datetime as _dt
    try:
        from supabase import create_client
    except Exception as e:
        raise RuntimeError(f"Supabase client not available: {e}. Ensure 'supabase' is in requirements.txt") from e

    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_KEY") or os.getenv("SUPABASE_SERVICE_KEY") or os.getenv("SUPABASE_ANON_KEY")
    if not url or not key:
        raise RuntimeError("Missing SUPABASE_URL and/or SUPABASE_KEY environment variables.")

    sb = create_client(url, key)
    resp = sb.table("rota_periods").select("start_date,end_date").eq("name", period_name).limit(1).execute()
    data = resp.data or []
    if not data:
        raise RuntimeError(f"No rota_periods row found with name={period_name!r}.")
    row = data[0]

    def _to_date(v):
        if v is None:
            return None
        if hasattr(v, "date") and not isinstance(v, str):
            try:
                return v.date()
            except Exception:
                pass
        if isinstance(v, str):
            s = v.strip()
            if not s:
                return None
            try:
                return _dt.fromisoformat(s.replace('Z','+00:00')).date()
            except Exception:
                from dateutil import parser as _parser
                return _parser.parse(s, dayfirst=True).date()
        raise RuntimeError(f"Unparseable date from Supabase: {v!r}")

    s = _to_date(row.get("start_date"))
    e = _to_date(row.get("end_date"))
    if not s or not e:
        raise RuntimeError(f"rota_periods row '{period_name}' missing start_date/end_date.")
    return s, e


@dataclass(frozen=True)
class Consultant:
    name: str
    cardiac: bool
    wte: float
    eligible_a: bool
    eligible_d: bool
    active: bool
    unique: str = ""
    unique_tags: frozenset[str] = frozenset()
    weekend_wte_cap: bool = False
    no_weekend: bool = False
    no_weekendab: bool = False
    no_d: bool = False


def _parse_unique_tags(val) -> frozenset[str]:
    if val is None:
        return frozenset()
    s = str(val).strip().lower()
    if not s:
        return frozenset()
    # Normalize common separators
    s = s.replace(";", ",").replace("\n", ",")
    parts = [p.strip() for p in s.split(",") if p.strip()]
    return frozenset(parts)

def _has_tag(tags: frozenset[str], needle: str) -> bool:
    needle = needle.strip().lower()
    return any(needle in t for t in tags)

def daterange(d0: date, d1: date) -> List[date]:
    out = []
    d = d0
    while d <= d1:
        out.append(d)
        d += timedelta(days=1)
    return out

def read_inputs(path: str) -> Tuple[date, date, List[Consultant], Dict[str, Set[date]], Set[date]]:
    wb = load_workbook(path, data_only=False)
    cfg = wb["Config"]

    # If dates were provided from Supabase, write them back into Config so downstream sheets stay consistent.
    if override_start is not None or override_end is not None:
        for r in range(1, 80):
            lab = str(cfg[f"A{r}"].value).strip() if cfg[f"A{r}"].value is not None else ""
            if lab == "CycleStartDate" and override_start is not None:
                cfg[f"B{r}"].value = override_start
            if lab == "CycleEndDate" and override_end is not None:
                cfg[f"B{r}"].value = override_end


    def get_cfg(label: str) -> Optional[date]:
        for r in range(1, 80):
            if str(cfg[f"A{r}"].value).strip() == label:
                return excel_date(cfg[f"B{r}"].value)
        raise ValueError(f"Config label not found: {label}")

    start = override_start or get_cfg("CycleStartDate")
    end = override_end or get_cfg("CycleEndDate")
    if start is None or end is None:
        raise ValueError(
            "CycleStartDate/CycleEndDate in the Config sheet are blank or not parseable as dates. "
            "Please ensure Config!B cell for those labels contains a real Excel date (not text), "
            "or an ISO date string like 2026-05-01."
        )


    cws = wb["Consultants"]
    consultants: List[Consultant] = []
    for r in range(2, 1000):
        nm = cws[f"A{r}"].value
        if not nm:
            continue
        tags = _parse_unique_tags(cws[f"H{r}"].value)
        consultants.append(Consultant(
            name=str(nm),
            cardiac=bool(cws[f"B{r}"].value),
            wte=float(cws[f"C{r}"].value or 0.0),
            eligible_a=bool(cws[f"D{r}"].value),
            eligible_d=bool(cws[f"E{r}"].value),
            active=bool(cws[f"F{r}"].value),
            unique=str(cws[f"H{r}"].value or ""),
            unique_tags=tags,
            weekend_wte_cap=_has_tag(tags, "weekendab = 1 wte"),
            no_weekendab=_has_tag(tags, "no weekendab"),
            no_weekend=_has_tag(tags, "no weekend"),
            no_d=_has_tag(tags, "no d"),
        ))
    consultants = [c for c in consultants if c.active]
    if not consultants:
        raise ValueError("No active consultants found.")

    lws = wb["Leave"]
    leave_map: Dict[str, Set[date]] = {c.name: set() for c in consultants}
    for r in range(2, 5000):
        nm = lws[f"A{r}"].value
        if not nm:
            continue
        if not bool(lws[f"E{r}"].value):
            continue
        s = excel_date(lws[f"B{r}"].value)
        e = excel_date(lws[f"C{r}"].value)
        if not s or not e:
            continue
        nm = str(nm)
        if nm not in leave_map:
            continue
        for d in daterange(s, e):
            leave_map[nm].add(d)

    bws = wb["BankHolidays"]
    bh: Set[date] = set()
    for r in range(2, 2000):
        d = excel_date(bws[f"A{r}"].value)
        if d:
            bh.add(d)

    return start, end, consultants, leave_map, bh

def solve(start: date, end: date, consultants: List[Consultant], leave: Dict[str, Set[date]], bank_holidays: Set[date],
          hard_no_consecutive_weekends: bool = True, hard_week_gap: bool = True, time_limit_s: int = 60, random_seed: int = 0) -> Dict:
    first_monday = start + timedelta(days=(7 - start.weekday()) % 7)
    weeks: List[date] = []
    d = first_monday
    while d <= end:
        weeks.append(d)
        d += timedelta(days=7)

    names = [c.name for c in consultants]
    N = len(names)
    cardiac = [c.cardiac for c in consultants]
    wte = [c.wte for c in consultants]
    eligible_a = [c.eligible_a for c in consultants]
    eligible_d = [c.eligible_d and (not c.no_d) for c in consultants]
    # Weekend eligibility: if no_weekend tag present, disallow all weekend blocks; if no_weekendab tag present, disallow WeekendAB only.
    weekend_allowed = [not c.no_weekend for c in consultants]
    weekendab_allowed = [not (c.no_weekend or c.no_weekendab) for c in consultants]
    # Weekend WTE for fairness: if tagged 'WeekendAB = 1 WTE', cap WTE at 1.0 for weekend distribution.
    weekend_wte = [min(c.wte, 1.0) if c.weekend_wte_cap else c.wte for c in consultants]

    model = cp_model.CpModel()
    block_types = ["AB1","AB2","DMonThu","WeekendAB","WeekendMixed"]
    x = {(w,b,i): model.NewBoolVar(f"x_{w}_{b}_{i}") for w in range(len(weeks)) for b in block_types for i in range(N)}

    for w_i in range(len(weeks)):
        for b in block_types:
            model.Add(sum(x[(w_i,b,i)] for i in range(N)) == 1)

    for w_i in range(len(weeks)):
        for i in range(N):
            if not eligible_a[i]:
                for b in ("AB1","AB2","WeekendAB","WeekendMixed"):
                    model.Add(x[(w_i,b,i)] == 0)
            if not eligible_d[i]:
                for b in ("DMonThu","WeekendMixed"):
                    model.Add(x[(w_i,b,i)] == 0)
            # Weekend restrictions from unique tags
            if not weekend_allowed[i]:
                for b in ("WeekendAB","WeekendMixed"):
                    model.Add(x[(w_i,b,i)] == 0)
            elif not weekendab_allowed[i]:
                model.Add(x[(w_i,"WeekendAB",i)] == 0)
            # No-D restriction from unique tags (redundant with eligible_d but kept explicit)
            if consultants[i].no_d:
                for b in ("DMonThu","WeekendMixed"):
                    model.Add(x[(w_i,b,i)] == 0)

    def block_days(week_monday: date, b: str) -> List[date]:
        if b == "AB1":
            return [week_monday + timedelta(days=k) for k in (0,1,2,3)]
        if b == "AB2":
            return [week_monday + timedelta(days=k) for k in (1,2,3,4)]
        if b == "DMonThu":
            return [week_monday + timedelta(days=k) for k in (0,1,2,3)]
        if b == "WeekendAB":
            return [week_monday + timedelta(days=k) for k in (4,5,6,7)]
        if b == "WeekendMixed":
            return [week_monday + timedelta(days=k) for k in (4,5,6)]
        raise ValueError(b)

    for w_i, wk in enumerate(weeks):
        for b in block_types:
            days = block_days(wk, b)
            for i,nm in enumerate(names):
                if any(d in leave.get(nm,set()) for d in days):
                    model.Add(x[(w_i,b,i)] == 0)

    for w_i in range(len(weeks)):
        for i in range(N):
            model.Add(sum(x[(w_i,b,i)] for b in block_types) <= 1)

    for i in range(N):
        for w_i in range(len(weeks)-1):
            wknd_this = x[(w_i,"WeekendAB",i)] + x[(w_i,"WeekendMixed",i)]
            wknd_next = x[(w_i+1,"WeekendAB",i)] + x[(w_i+1,"WeekendMixed",i)]
            if hard_no_consecutive_weekends:
                model.Add(wknd_this + wknd_next <= 1)

    if hard_week_gap:
        for i in range(N):
            for w_i in range(len(weeks)-1):
                any_this = sum(x[(w_i,b,i)] for b in block_types)
                any_next = sum(x[(w_i+1,b,i)] for b in block_types)
                model.Add(any_this + any_next <= 1)

    # Cardiac XOR weekdays Mon-Fri
    for w_i in range(len(weeks)):
        for day in range(5):
            # D cardiac: Mon-Thu from DMonThu, Fri from WeekendMixed
            if day <= 3:
                D_c = sum(x[(w_i,"DMonThu",i)] * (1 if cardiac[i] else 0) for i in range(N))
            else:
                D_c = sum(x[(w_i,"WeekendMixed",i)] * (1 if cardiac[i] else 0) for i in range(N))
            # A cardiac: Mon/Wed AB1, Tue/Thu AB2, Fri WeekendAB
            if day in (0,2):
                A_c = sum(x[(w_i,"AB1",i)] * (1 if cardiac[i] else 0) for i in range(N))
            elif day in (1,3):
                A_c = sum(x[(w_i,"AB2",i)] * (1 if cardiac[i] else 0) for i in range(N))
            else:
                A_c = sum(x[(w_i,"WeekendAB",i)] * (1 if cardiac[i] else 0) for i in range(N))
            model.Add(A_c + D_c == 1)
    # -----------------------------
    
    # -----------------------------
    # Preferred shifts (soft constraints in objective)
    #
    # We accept either of these header sets in the sheet:
    #   (A) name | start_date | end_date | kind | shift_type | weight | notes
    #   (B) consultant_name | start_date | end_date | pref_kind | shift_type | weight | notes | updated_at
    #
    # shift_type accepted from your UI:
    #   A -> AB1, B -> AB2, D -> DMonThu, Weekend -> WeekendAB + WeekendMixed
    #
    # Optional "avoid" semantics: if pref_kind or notes contains 'avoid'/'do not'/'no'
    # -----------------------------
    pref_items = []
    try:
        if "preferred_shifts" in wb.sheetnames:
            pws = wb["preferred_shifts"]

            # Read headers
            headers = [str(c.value).strip() if c.value is not None else "" for c in next(pws.iter_rows(min_row=1, max_row=1))]
            hmap = {h.lower(): idx for idx, h in enumerate(headers) if h}

            def get(row, *names):
                for nm in names:
                    j = hmap.get(nm.lower())
                    if j is not None and j < len(row):
                        return row[j]
                return None

            for row in pws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                raw_name = get(row, "name", "consultant_name")
                if not raw_name:
                    continue
                name = str(raw_name).strip()
                if not name or name not in name_to_i:
                    continue

                sd = excel_date(get(row, "start_date"))
                ed = excel_date(get(row, "end_date"))
                if not sd or not ed:
                    continue

                kind = get(row, "kind", "pref_kind")
                notes = get(row, "notes")
                stype = get(row, "shift_type")
                weight = get(row, "weight")

                kind_s = str(kind).strip().lower() if kind is not None else ""
                notes_s = str(notes).strip().lower() if notes is not None else ""
                stype_s = str(stype).strip() if stype is not None else ""

                try:
                    w = int(weight) if weight is not None else 3
                except Exception:
                    w = 3

                # Determine sign: prefer => reward, avoid => penalty
                avoid_tokens = ["avoid", "do not", "dont", "don't", "no "]
                is_avoid = any(t in kind_s for t in avoid_tokens) or any(t in notes_s for t in avoid_tokens)
                sign = -1 if is_avoid else 1

                # Map shift types to block types
                blocks = []
                st_u = stype_s.strip().upper()
                if st_u == "A":
                    blocks = ["AB1"]
                elif st_u == "B":
                    blocks = ["AB2"]
                elif st_u == "D":
                    blocks = ["DMonThu"]
                elif st_u in ["WEEKEND", "WKND", "W/END"]:
                    blocks = ["WeekendAB", "WeekendMixed"]
                elif stype_s in block_types:
                    blocks = [stype_s]
                else:
                    st_lower = stype_s.lower()
                    if st_lower in ["ab", "a/b", "abab", "ab1+ab2", "weekday_ab"]:
                        blocks = ["AB1", "AB2"]
                    elif st_lower in ["weekend", "wknd", "weekends"]:
                        blocks = ["WeekendAB", "WeekendMixed"]
                    elif st_lower in ["d", "dmonthu"]:
                        blocks = ["DMonThu"]
                    else:
                        # If blank/unknown, treat as weekend preference when date range includes Sat/Sun, else AB
                        has_weekend = any((sd + timedelta(days=k)).weekday() >= 5 for k in range((ed - sd).days + 1))
                        blocks = ["WeekendAB", "WeekendMixed"] if has_weekend else ["AB1", "AB2"]

                # Expand date range into rota weeks (Monday dates)
                def week_monday(d: date) -> date:
                    return d - timedelta(days=d.weekday())

                w_start = week_monday(sd)
                w_end = week_monday(ed)
                wks = []
                w_cur = w_start
                while w_cur <= w_end:
                    if w_cur in week_index:
                        wks.append(w_cur)
                    w_cur += timedelta(days=7)

                if wks:
                    pref_items.append({"i": name_to_i[name], "weeks": wks, "blocks": blocks, "sign": sign, "weight": max(1, w)})
    except Exception:
        # Preferences are optional; ignore parsing issues
        pref_items = []

# Objective: WTE-weighted fairness (total, BH, weekends)
    block_weight = {"AB1":4, "AB2":4, "DMonThu":4, "WeekendAB":4, "WeekendMixed":3}
    total_duty = [model.NewIntVar(0, 20000, f"total_{i}") for i in range(N)]
    for i in range(N):
        model.Add(total_duty[i] == sum(x[(w_i,b,i)] * block_weight[b] for w_i in range(len(weeks)) for b in block_types))

    total_all = sum(block_weight[b] for b in block_types) * len(weeks)
    sum_wte = sum(wte) if sum(wte) > 0 else 1.0
    SCALE = 1000

    expected = [int(round(total_all * (wte[i]/sum_wte) * SCALE)) for i in range(N)]
    actualS = [model.NewIntVar(0, 10_000_000, f"aS_{i}") for i in range(N)]
    devT = [model.NewIntVar(0, 10_000_000, f"devT_{i}") for i in range(N)]
    for i in range(N):
        model.Add(actualS[i] == total_duty[i] * SCALE)
        model.AddAbsEquality(devT[i], actualS[i] - expected[i])

    # BH proxy counts
    bh_count = {}
    for w_i, wk in enumerate(weeks):
        for b in block_types:
            bh_count[(w_i,b)] = sum(1 for d in block_days(wk,b) if d in bank_holidays)
    bh_duty = [model.NewIntVar(0, 20000, f"bh_{i}") for i in range(N)]
    for i in range(N):
        model.Add(bh_duty[i] == sum(x[(w_i,b,i)] * bh_count[(w_i,b)] for w_i in range(len(weeks)) for b in block_types))

    bh_all = sum(bh_count[(w_i,b)] for w_i in range(len(weeks)) for b in block_types)
    expected_bh = [int(round(bh_all * (wte[i]/sum_wte) * SCALE)) for i in range(N)]
    bhS = [model.NewIntVar(0, 10_000_000, f"bhS_{i}") for i in range(N)]
    devBH = [model.NewIntVar(0, 10_000_000, f"devBH_{i}") for i in range(N)]
    for i in range(N):
        model.Add(bhS[i] == bh_duty[i] * SCALE)
        model.AddAbsEquality(devBH[i], bhS[i] - expected_bh[i])

    weekend_blocks = [model.NewIntVar(0, 20000, f"wknd_{i}") for i in range(N)]
    for i in range(N):
        model.Add(weekend_blocks[i] == sum(x[(w_i,"WeekendAB",i)] + x[(w_i,"WeekendMixed",i)] for w_i in range(len(weeks))))
    weekend_all = 2 * len(weeks)
    # Distribute weekend burden only among those eligible for weekend blocks.
    sum_wte_weekend = sum(weekend_wte[i] for i in range(N) if weekend_allowed[i])
    sum_wte_weekend = sum_wte_weekend if sum_wte_weekend > 0 else 1.0
    expected_w = [0 for _ in range(N)]
    for i in range(N):
        if weekend_allowed[i]:
            expected_w[i] = int(round(weekend_all * (weekend_wte[i]/sum_wte_weekend) * SCALE))

    wkS = [model.NewIntVar(0, 10_000_000, f"wkS_{i}") for i in range(N)]
    devW = [model.NewIntVar(0, 10_000_000, f"devW_{i}") for i in range(N)]
    for i in range(N):
        model.Add(wkS[i] == weekend_blocks[i] * SCALE)
        model.AddAbsEquality(devW[i], wkS[i] - expected_w[i])

    model.Minimize(sum(devT) + 3*sum(devBH) + 2*sum(devW))

    solver = cp_model.CpSolver()
    solver.parameters.random_seed = int(random_seed)
    solver.parameters.randomize_search = True
    solver.parameters.max_time_in_seconds = float(time_limit_s)
    solver.parameters.num_search_workers = 8

    status = solver.Solve(model)
    status_name = solver.StatusName(status)
    objective = solver.ObjectiveValue() if status in (cp_model.OPTIMAL, cp_model.FEASIBLE) else None

    sol = {"status": status_name, "objective": objective, "weeks": weeks, "assignments": {wk:{} for wk in weeks}}
    sol["pref_score"] = None

    if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        for w_i, wk in enumerate(weeks):
            for b in block_types:
                for i in range(N):
                    if solver.Value(x[(w_i,b,i)]) == 1:
                        sol["assignments"][wk][b] = names[i]
                        break
    if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        # Simple preference score: sum(weight) for satisfied prefers minus satisfied avoids
        score = 0
        for p in pref_items:
            i = p["i"]
            name = names[i]
            for wk in p["weeks"]:
                asg = sol["assignments"].get(wk, {})
                hit = any(asg.get(b, "") == name for b in p["blocks"])
                if hit:
                    score += (p["weight"] if p["sign"] == 1 else -p["weight"])
        sol["pref_score"] = int(score)
    return sol

def export_to_excel(input_path: str, output_path: str, sol: Dict, override_start: Optional[date] = None, override_end: Optional[date] = None):
    wb = load_workbook(input_path)

    # -----------------------------
    # Read preferred_shifts (currently recorded for future scoring; not enforced in this solver version)
    # -----------------------------
    preferred = []
    if "preferred_shifts" in wb.sheetnames:
        pws = wb["preferred_shifts"]
        for row in pws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            # Expected columns: name, start_date, end_date, kind, shift_type, weight, notes
            name, sd, ed, kind, shift_type, weight, notes = (list(row) + [None]*7)[:7]
            try:
                weight = int(weight) if weight is not None else 3
            except Exception:
                weight = 3
            preferred.append(
                {"name": str(name), "start_date": sd, "end_date": ed, "shift": shift_type, "weight": weight}
            )

    wa = wb["WeekAssignments"]
    rota = wb["Rota"]
    dash = wb["Dashboard"]
    cfg = wb["Config"]

    # If dates were provided from Supabase, write them back into Config so downstream sheets stay consistent.
    if override_start is not None or override_end is not None:
        for r in range(1, 80):
            lab = str(cfg[f"A{r}"].value).strip() if cfg[f"A{r}"].value is not None else ""
            if lab == "CycleStartDate" and override_start is not None:
                cfg[f"B{r}"].value = override_start
            if lab == "CycleEndDate" and override_end is not None:
                cfg[f"B{r}"].value = override_end

    cons = wb["Consultants"]
    leave_ws = wb["Leave"]
    bh_ws = wb["BankHolidays"]

    def get_cfg(label: str):
        for r in range(1, 80):
            if str(cfg[f"A{r}"].value).strip() == label:
                return cfg[f"B{r}"].value
        return None

    start = excel_date(get_cfg("CycleStartDate"))
    end = excel_date(get_cfg("CycleEndDate"))

    prev_A_for_start = str(get_cfg("A_Consultant_DayBeforeStart") or "")

    cardiac = {}
    wte = {}
    for r in range(2, 1000):
        nm = cons[f"A{r}"].value
        if not nm:
            continue
        nm = str(nm)
        cardiac[nm] = bool(cons[f"B{r}"].value)
        wte[nm] = float(cons[f"C{r}"].value or 0.0)

    leave_map = {}
    for r in range(2, 5000):
        nm = leave_ws[f"A{r}"].value
        if not nm:
            continue
        if not bool(leave_ws[f"E{r}"].value):
            continue
        s = excel_date(leave_ws[f"B{r}"].value)
        e = excel_date(leave_ws[f"C{r}"].value)
        if not s or not e:
            continue
        nm = str(nm)
        leave_map.setdefault(nm, set())
        d = s
        while d <= e:
            leave_map[nm].add(d)
            d += timedelta(days=1)

    bh_set = set()
    for r in range(2, 2000):
        d = excel_date(bh_ws[f"A{r}"].value)
        if d:
            bh_set.add(d)

    # Clear WeekAssignments
    for r in range(2, wa.max_row + 1):
        for c in range(1, 9):
            wa.cell(r, c).value = None

    weeks = sol["weeks"]
    for r_i, wk in enumerate(weeks, start=2):
        wa.cell(r_i, 1).value = wk
        wa.cell(r_i, 2).value = sol["assignments"][wk].get("AB1", "")
        wa.cell(r_i, 3).value = sol["assignments"][wk].get("AB2", "")
        wa.cell(r_i, 4).value = sol["assignments"][wk].get("DMonThu", "")
        wa.cell(r_i, 5).value = sol["assignments"][wk].get("WeekendAB", "")
        wa.cell(r_i, 6).value = sol["assignments"][wk].get("WeekendMixed", "")
        wa.cell(r_i, 7).value = sol.get("status", "")
        wa.cell(r_i, 8).value = sol.get("objective", "")

    wk_map = {wk: sol["assignments"][wk] for wk in weeks}

    def week_monday(d: date) -> date:
        return d - timedelta(days=d.weekday())

    # Clear Rota
    for r in range(2, rota.max_row + 1):
        for c in range(1, 7):
            rota.cell(r, c).value = None

    all_days = daterange(start, end)
    prev_A = None
    for row_i, d in enumerate(all_days, start=2):
        dow = d.weekday()  # Mon=0..Sun=6
        wk = week_monday(d)
        asg = wk_map.get(wk, {})

        if dow in (0, 2):  # Mon/Wed
            A = asg.get("AB1", "")
        elif dow in (1, 3):  # Tue/Thu
            A = asg.get("AB2", "")
        elif dow == 4:  # Fri
            A = asg.get("WeekendAB", "")
        elif dow == 5:  # Sat
            A = asg.get("WeekendMixed", "")
        else:  # Sun
            A = asg.get("WeekendAB", "")

        if d == start:
            B = prev_A_for_start
        else:
            B = prev_A or ""

        if dow <= 3:
            D = asg.get("DMonThu", "")
        elif dow == 4:
            D = asg.get("WeekendMixed", "")
        else:
            D = ""  # No D at weekends

        flags = []
        if not A:
            flags.append("MISSING_A")
        if not B:
            flags.append("MISSING_B")
        if dow <= 4 and not D:
            flags.append("MISSING_D")
        if dow >= 5 and D:
            flags.append("D_SHOULD_BE_BLANK_WEEKEND")

        if A and d in leave_map.get(A, set()):
            flags.append("A_ON_LEAVE")
        if B and d in leave_map.get(B, set()):
            flags.append("B_ON_LEAVE")
        if D and d in leave_map.get(D, set()):
            flags.append("D_ON_LEAVE")

        if dow <= 4:
            a_c = bool(cardiac.get(A, False))
            d_c = bool(cardiac.get(D, False))
            if (a_c + d_c) != 1:
                flags.append("CARDIAC_XOR_BREACH")

        if d in bh_set:
            flags.append("BANK_HOLIDAY")

        rota.cell(row_i, 1).value = d
        rota.cell(row_i, 2).value = d.strftime("%a")
        rota.cell(row_i, 3).value = A
        rota.cell(row_i, 4).value = B
        rota.cell(row_i, 5).value = D
        rota.cell(row_i, 6).value = ",".join(flags)

        prev_A = A

    # Dashboard (values)
    for r in range(2, dash.max_row + 1):
        for c in range(1, 14):
            dash.cell(r, c).value = None

    counts = {nm: {"A": 0, "B": 0, "D": 0, "BH": 0, "wknd": 0, "consec_wknd": 0} for nm in cardiac.keys()}
    weekend_by_cons = {nm: [] for nm in cardiac.keys()}
    for wk in weeks:
        for b in ("WeekendAB", "WeekendMixed"):
            nm = wk_map[wk].get(b, "")
            if nm:
                weekend_by_cons.setdefault(nm, []).append(wk)

    for nm, wks in weekend_by_cons.items():
        wks = sorted(wks)
        for i in range(len(wks) - 1):
            if (wks[i + 1] - wks[i]).days == 7:
                counts[nm]["consec_wknd"] += 1
        counts[nm]["wknd"] = len(wks)

    for row_i, d in enumerate(all_days, start=2):
        A = rota.cell(row_i, 3).value or ""
        B = rota.cell(row_i, 4).value or ""
        Dv = rota.cell(row_i, 5).value or ""
        is_bh = "BANK_HOLIDAY" in (rota.cell(row_i, 6).value or "")
        if A in counts:
            counts[A]["A"] += 1
        if B in counts:
            counts[B]["B"] += 1
        if Dv in counts and d.weekday() <= 4:
            counts[Dv]["D"] += 1
        if is_bh:
            if A in counts:
                counts[A]["BH"] += 1
            if B in counts:
                counts[B]["BH"] += 1
            if Dv in counts and d.weekday() <= 4:
                counts[Dv]["BH"] += 1

    total_all = sum(v["A"] + v["B"] + v["D"] for v in counts.values())
    total_bh = sum(v["BH"] for v in counts.values())
    sum_wte = sum(wte.values()) if wte else 1.0

    r = 2
    for nm in sorted(counts.keys()):
        A_cnt = counts[nm]["A"]
        B_cnt = counts[nm]["B"]
        D_cnt = counts[nm]["D"]
        tot = A_cnt + B_cnt + D_cnt
        exp = total_all * (wte.get(nm, 0.0) / sum_wte)
        delta = tot - exp
        bh_cnt = counts[nm]["BH"]
        bh_exp = total_bh * (wte.get(nm, 0.0) / sum_wte)
        bh_delta = bh_cnt - bh_exp

        dash.cell(r, 1).value = nm
        dash.cell(r, 2).value = wte.get(nm, 0.0)
        dash.cell(r, 3).value = A_cnt
        dash.cell(r, 4).value = B_cnt
        dash.cell(r, 5).value = D_cnt
        dash.cell(r, 6).value = tot
        dash.cell(r, 7).value = exp
        dash.cell(r, 8).value = delta
        dash.cell(r, 9).value = bh_cnt
        dash.cell(r, 10).value = bh_exp
        dash.cell(r, 11).value = bh_delta
        dash.cell(r, 12).value = counts[nm]["wknd"]
        dash.cell(r, 13).value = counts[nm]["consec_wknd"]
        r += 1

    wb.save(output_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--time_limit", type=int, default=60)
    ap.add_argument("--seed", type=int, default=0)
    ap.add_argument("--period_name", type=str, default="")  # reads dates from Supabase rota_periods.name
    ap.add_argument("--no_hard_week_gap", action="store_true")
    ap.add_argument("--no_hard_no_consec_weekends", action="store_true")
    args = ap.parse_args()

    override_start = None
    override_end = None
    if args.period_name:
        override_start, override_end = fetch_period_dates_from_supabase(args.period_name)

    try:
        start, end, consultants, leave, bh = read_inputs(args.input, override_start=override_start, override_end=override_end)
    except Exception as e:
        raise SystemExit(f"Input workbook error: {e}")
    sol = solve(
        start, end, consultants, leave, bh,
        hard_no_consecutive_weekends=not args.no_hard_no_consec_weekends,
        hard_week_gap=not args.no_hard_week_gap,
        time_limit_s=args.time_limit,
        random_seed=args.seed,
    )
    print(f"Status: {sol['status']}  Objective: {sol.get('objective')}")
    export_to_excel(args.input, args.output, sol, override_start=override_start, override_end=override_end)
    print(f"Wrote {args.output}")

if __name__ == "__main__":
    main()


# -----------------------------
# PREFERRED SHIFTS OBJECTIVE (SOFT)
# -----------------------------
# Assumes decision variables exist for assignments, e.g.:
#   x[(consultant, date, shift)] = BoolVar
#
# For each preferred shift request:
#   maximize weight * x[(name, d, shift)] for d in [start_date..end_date]
#
# Example (pseudo-code):
# pref_terms = []
# for p in preferred:
#     for d in daterange(p['start_date'], p['end_date']):
#         key = (p['name'], d, p['shift'])
#         if key in x:
#             pref_terms.append(p['weight'] * x[key])
# model.Maximize(sum(pref_terms) + existing_objective)
#
# NOTE:
# If you already minimize penalties, instead add:
#   model.Minimize(existing_penalty - sum(pref_terms))
#
# Tune scaling factor if needed (e.g. p['weight'] * 10).
