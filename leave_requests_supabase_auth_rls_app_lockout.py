import streamlit as st
from datetime import date, datetime
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
from supabase import create_client, Client

st.set_page_config(page_title="Rota Leave Requests", layout="wide")

# -----------------------------
# Secrets
# -----------------------------
SUPABASE_URL = st.secrets.get("SUPABASE_URL", "")
SUPABASE_ANON_KEY = st.secrets.get("SUPABASE_ANON_KEY", "")  # public key
ALLOWED_EMAIL_DOMAIN = st.secrets.get("ALLOWED_EMAIL_DOMAIN", "")  # optional

if not SUPABASE_URL or not SUPABASE_ANON_KEY:
    st.error("Missing SUPABASE_URL or SUPABASE_ANON_KEY in Streamlit secrets.")
    st.stop()

ALLOWED_TYPES = ["Annual", "Study", "NOC"]

# -----------------------------
# Supabase client helpers
# -----------------------------
def base_client() -> Client:
    return create_client(SUPABASE_URL, SUPABASE_ANON_KEY)

def authed_client(access_token: str) -> Client:
    c = base_client()
    c.postgrest.auth(access_token)  # enforce RLS
    return c

def validate_dates(s: date, e: date) -> str | None:
    if e < s:
        return "Date to cannot be earlier than Date from."
    return None

def overlap(a_start: date, a_end: date, b_start: date, b_end: date) -> bool:
    return a_start <= b_end and b_start <= a_end

# -----------------------------
# Auth UI
# -----------------------------
st.title("Rota Leave Requests")
st.write(
    "Supabase Auth + Row Level Security (RLS), with **automatic lockout** after a rota period is published. "
    "Consultants can only manage requests that do not overlap a published period."
)

if "sb_session" not in st.session_state:
    st.session_state["sb_session"] = None

with st.sidebar:
    st.header("Sign in")
    email = st.text_input("Email", value="").strip().lower()
    password = st.text_input("Password", type="password")

    if ALLOWED_EMAIL_DOMAIN and email and not email.endswith("@" + ALLOWED_EMAIL_DOMAIN):
        st.warning(f"Email must end with @{ALLOWED_EMAIL_DOMAIN}.")

    c = base_client()

    col1, col2 = st.columns(2)
    with col1:
        if st.button("Sign in", use_container_width=True):
            if not email or not password:
                st.error("Enter email and password.")
            elif ALLOWED_EMAIL_DOMAIN and not email.endswith("@" + ALLOWED_EMAIL_DOMAIN):
                st.error("Email domain not permitted.")
            else:
                try:
                    res = c.auth.sign_in_with_password({"email": email, "password": password})
                    st.session_state["sb_session"] = res.session.model_dump() if res.session else None
                    st.rerun()
                except Exception:
                    st.error("Sign-in failed. Check credentials and confirmation status.")

    with col2:
        if st.button("Sign up", use_container_width=True):
            if not email or not password:
                st.error("Enter email and password.")
            elif ALLOWED_EMAIL_DOMAIN and not email.endswith("@" + ALLOWED_EMAIL_DOMAIN):
                st.error("Email domain not permitted.")
            else:
                try:
                    _ = c.auth.sign_up({"email": email, "password": password})
                    st.success("Sign-up created. If email confirmation is enabled, confirm via email then sign in.")
                except Exception:
                    st.error("Sign-up failed. Email may already exist or password may be too weak.")

    if st.session_state["sb_session"]:
        if st.button("Sign out", use_container_width=True):
            st.session_state["sb_session"] = None
            st.rerun()

sess = st.session_state["sb_session"]
if not sess:
    st.info("Please sign in to continue.")
    st.stop()

access_token = sess["access_token"]
user_id = sess["user"]["id"]
user_email = sess["user"]["email"]

db = authed_client(access_token)

# -----------------------------
# Admin detection
# -----------------------------
def is_rota_admin() -> bool:
    try:
        r = db.table("rota_admins").select("user_id").limit(1).execute()
        return bool(r.data)
    except Exception:
        return False

is_admin = is_rota_admin()

with st.sidebar:
    st.markdown("---")
    st.write(f"Signed in as: {user_email}")
    st.write("Role: Rota admin" if is_admin else "Role: Consultant")

# -----------------------------
# Published periods
# -----------------------------
def fetch_periods() -> pd.DataFrame:
    resp = db.table("rota_periods").select("*").order("start_date").execute()
    data = resp.data or []
    dfp = pd.DataFrame(data)
    if not dfp.empty:
        dfp["start_date"] = pd.to_datetime(dfp["start_date"]).dt.date
        dfp["end_date"] = pd.to_datetime(dfp["end_date"]).dt.date
        dfp["published_at"] = pd.to_datetime(dfp["published_at"], errors="coerce")
        dfp["created_at"] = pd.to_datetime(dfp["created_at"], errors="coerce")
    return dfp

periods = fetch_periods()

def is_locked_for_range(s: date, e: date) -> bool:
    if periods.empty:
        return False
    pubs = periods[periods["is_published"] == True]
    if pubs.empty:
        return False
    for _, p in pubs.iterrows():
        if overlap(s, e, p["start_date"], p["end_date"]):
            return True
    return False

st.subheader("0) Published rota periods (lockout)")
if periods.empty:
    st.info("No rota periods configured yet.")
else:
    st.dataframe(periods[["id","name","start_date","end_date","is_published","published_at"]],
                 use_container_width=True, hide_index=True)
    if not periods[periods["is_published"] == True].empty:
        st.caption("Requests overlapping published periods cannot be created/edited/deleted by consultants.")

# -----------------------------
# Data access
# -----------------------------
def fetch_leave_requests() -> pd.DataFrame:
    resp = db.table("leave_requests").select("*").order("start_date").execute()
    data = resp.data or []
    df = pd.DataFrame(data)
    if not df.empty:
        df["start_date"] = pd.to_datetime(df["start_date"]).dt.date
        df["end_date"] = pd.to_datetime(df["end_date"]).dt.date
        df["created_at"] = pd.to_datetime(df["created_at"])
        df["updated_at"] = pd.to_datetime(df["updated_at"])
    return df

df = fetch_leave_requests()

# -----------------------------
# Add request
# -----------------------------
st.subheader("1) Submit a leave request")
with st.form("add_form"):
    c1, c2, c3 = st.columns([2, 1, 1])
    with c1:
        consultant_name = st.text_input("Consultant name (as it should appear on the rota)", value="")
        leave_type = st.selectbox("Leave type", options=ALLOWED_TYPES)
        notes = st.text_input("Notes (optional)", value="")
    with c2:
        start_date = st.date_input("Date from", value=date.today())
    with c3:
        end_date = st.date_input("Date to", value=date.today())
    submitted = st.form_submit_button("Submit request")

if submitted:
    err = validate_dates(start_date, end_date)
    if err:
        st.error(err)
    elif not consultant_name.strip():
        st.error("Consultant name is required.")
    else:
        if (not is_admin) and is_locked_for_range(start_date, end_date):
            st.error("This rota period has been published. You cannot submit leave that overlaps a published period.")
        else:
            try:
                db.table("leave_requests").insert({
                    "consultant_name": consultant_name.strip(),
                    "requester_id": user_id,
                    "requester_email": user_email,
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat(),
                    "leave_type": leave_type,
                    "approved": False,
                    "notes": notes.strip()
                }).execute()
                st.success("Request submitted (pending approval).")
                st.rerun()
            except Exception:
                st.error("Insert failed (likely locked by published period or RLS not applied).")

# -----------------------------
# View requests
# -----------------------------
st.subheader("2) Your requests")
if df.empty:
    st.info("No requests found.")
else:
    view_cols = ["id", "consultant_name", "start_date", "end_date", "leave_type", "approved", "notes", "updated_at"]
    st.dataframe(df[view_cols], use_container_width=True, hide_index=True)

# -----------------------------
# Edit/delete pending
# -----------------------------
st.subheader("3) Edit or delete a pending request")
pending = df[df["approved"] == False].copy() if not df.empty else pd.DataFrame()
if pending.empty:
    st.info("No pending requests available to edit/delete.")
else:
    selected_id = st.selectbox("Select request", options=pending["id"].tolist())
    row = pending[pending["id"] == selected_id].iloc[0]

    locked_row = (not is_admin) and is_locked_for_range(row["start_date"], row["end_date"])
    if locked_row:
        st.warning("This request overlaps a published period and is locked. Only an admin can modify it.")

    with st.form("edit_form"):
        e1, e2, e3 = st.columns([2, 1, 1])
        with e1:
            new_name = st.text_input("Consultant name", value=row["consultant_name"], disabled=locked_row)
            new_type = st.selectbox("Leave type", options=ALLOWED_TYPES,
                                    index=ALLOWED_TYPES.index(row["leave_type"]), disabled=locked_row)
            new_notes = st.text_input("Notes", value=row.get("notes") or "", disabled=locked_row)
        with e2:
            new_start = st.date_input("Date from", value=row["start_date"], disabled=locked_row)
        with e3:
            new_end = st.date_input("Date to", value=row["end_date"], disabled=locked_row)

        csave, cdel = st.columns([1, 1])
        with csave:
            save = st.form_submit_button("Save changes", disabled=locked_row)
        with cdel:
            delete = st.form_submit_button("Delete request", disabled=locked_row)

    if save:
        err = validate_dates(new_start, new_end)
        if err:
            st.error(err)
        elif not new_name.strip():
            st.error("Consultant name is required.")
        else:
            if (not is_admin) and is_locked_for_range(new_start, new_end):
                st.error("This overlaps a published period and cannot be modified.")
            else:
                try:
                    db.table("leave_requests").update({
                        "consultant_name": new_name.strip(),
                        "start_date": new_start.isoformat(),
                        "end_date": new_end.isoformat(),
                        "leave_type": new_type,
                        "notes": new_notes.strip()
                    }).eq("id", selected_id).execute()
                    st.success("Updated.")
                    st.rerun()
                except Exception:
                    st.error("Update failed (likely locked/published, or RLS not installed).")

    if delete:
        try:
            db.table("leave_requests").delete().eq("id", selected_id).execute()
            st.success("Deleted.")
            st.rerun()
        except Exception:
            st.error("Delete failed (likely locked/published, or RLS not installed).")

# -----------------------------
# Admin actions: manage periods, approve, compile
# -----------------------------
st.subheader("4) Admin actions")
if not is_admin:
    st.info("Admin actions are available to rota administrators only.")
    st.stop()

st.markdown("#### Manage rota periods (publish/unpublish)")
with st.expander("Create a new rota period", expanded=False):
    with st.form("create_period"):
        p1, p2, p3, p4 = st.columns([2, 1, 1, 1])
        with p1:
            pname = st.text_input("Period name", value="Nov–May 20XX/20XX")
        with p2:
            pstart = st.date_input("Start date", value=date.today(), key="pstart")
        with p3:
            pend = st.date_input("End date", value=date.today(), key="pend")
        with p4:
            publish_now = st.checkbox("Publish now", value=False)
        create_btn = st.form_submit_button("Create period")
    if create_btn:
        err = validate_dates(pstart, pend)
        if err:
            st.error(err)
        else:
            payload = {
                "name": (pname.strip() or f"{pstart} to {pend}"),
                "start_date": pstart.isoformat(),
                "end_date": pend.isoformat(),
                "is_published": bool(publish_now),
                "published_at": datetime.utcnow().replace(microsecond=0).isoformat() + "Z" if publish_now else None,
                "created_by": user_id
            }
            try:
                db.table("rota_periods").insert(payload).execute()
                st.success("Period created.")
                st.rerun()
            except Exception:
                st.error("Create failed. Confirm rota_periods table + RLS policies are installed.")

periods = fetch_periods()
if not periods.empty:
    sel_pid = st.selectbox("Select period to manage", options=periods["id"].tolist(), key="pid")
    prow = periods[periods["id"] == sel_pid].iloc[0]
    cA, cB, cC = st.columns([1, 1, 2])
    with cA:
        if st.button("Publish selected", use_container_width=True, disabled=bool(prow["is_published"])):
            db.table("rota_periods").update({
                "is_published": True,
                "published_at": datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
            }).eq("id", sel_pid).execute()
            st.success("Published.")
            st.rerun()
    with cB:
        if st.button("Unpublish selected", use_container_width=True, disabled=not bool(prow["is_published"])):
            db.table("rota_periods").update({
                "is_published": False,
                "published_at": None
            }).eq("id", sel_pid).execute()
            st.success("Unpublished.")
            st.rerun()
    with cC:
        st.caption("Publishing activates lockout for consultants on overlapping dates (create/update/delete).")

st.markdown("---")
st.markdown("#### Approvals")
df_admin = fetch_leave_requests()
pending_admin = df_admin[df_admin["approved"] == False].copy() if not df_admin.empty else pd.DataFrame()
if pending_admin.empty:
    st.write("No pending requests.")
else:
    st.dataframe(pending_admin[["id","consultant_name","requester_email","start_date","end_date","leave_type","notes","created_at"]],
                 use_container_width=True, hide_index=True)
    approve_id = st.selectbox("Select pending request ID", options=pending_admin["id"].tolist(), key="appr")
    colA, colB = st.columns([1, 1])
    with colA:
        if st.button("Approve selected", use_container_width=True):
            db.table("leave_requests").update({"approved": True}).eq("id", approve_id).execute()
            st.success("Approved.")
            st.rerun()
    with colB:
        if st.button("Reject (delete) selected", use_container_width=True):
            db.table("leave_requests").delete().eq("id", approve_id).execute()
            st.success("Rejected (deleted).")
            st.rerun()

st.markdown("---")
st.markdown("#### Compile approved leave into Excel workbook (download)")
template = st.file_uploader("Upload rota workbook (.xlsx)", type=["xlsx"], key="tmpl")
only_approved = st.checkbox("Export approved only", value=True)
if template is None:
    st.info("Upload a workbook to enable compilation.")
else:
    if st.button("Compile and generate workbook"):
        export_df = fetch_leave_requests()
        if only_approved:
            export_df = export_df[export_df["approved"] == True].copy()

        wb = load_workbook(BytesIO(template.getvalue()))
        if "Leave" not in wb.sheetnames:
            st.error("Workbook must contain a sheet named 'Leave'.")
            st.stop()
        lws = wb["Leave"]

        for r in range(2, 5000):
            if lws[f"A{r}"].value in (None, ""):
                break
            for col in ("A","B","C","D","E"):
                lws[f"{col}{r}"].value = None

        r = 2
        if not export_df.empty:
            export_df = export_df.sort_values(["start_date","consultant_name"], na_position="last")
            for _, rec in export_df.iterrows():
                lws[f"A{r}"].value = rec["consultant_name"]
                lws[f"B{r}"].value = rec["start_date"]
                lws[f"C{r}"].value = rec["end_date"]
                lws[f"D{r}"].value = rec["leave_type"]
                lws[f"E{r}"].value = bool(rec["approved"])
                r += 1

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        st.download_button(
            "Download compiled workbook",
            data=out,
            file_name="Rota_Master_WITH_Leave.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
