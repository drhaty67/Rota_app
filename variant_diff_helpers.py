"""
Variant diff helpers for Streamlit rota app.

Goal: provide a "visual difference between variants" UI by comparing two solved XLSX workbooks
cell-by-cell for a chosen sheet (default: 'Rota').

Usage in Streamlit:
    from variant_diff_helpers import common_sheets, diff_sheet, diff_summary

    sheets = common_sheets(bytes_a, bytes_b)
    sheet = st.selectbox("Sheet to compare", sheets, index=sheets.index("Rota") if "Rota" in sheets else 0)
    diffs = diff_sheet(bytes_a, bytes_b, sheet, max_changes=5000)
    summary = diff_summary(diffs)
"""

from __future__ import annotations

from io import BytesIO
from typing import List, Tuple
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

def _norm(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        # avoid 1 vs 1.0 noise
        if abs(v - round(v)) < 1e-9:
            return str(int(round(v)))
        return f"{v:.10g}"
    return str(v).strip()

def common_sheets(xlsx_a: bytes, xlsx_b: bytes) -> List[str]:
    wb_a = openpyxl.load_workbook(BytesIO(xlsx_a), data_only=True)
    wb_b = openpyxl.load_workbook(BytesIO(xlsx_b), data_only=True)
    return [s for s in wb_a.sheetnames if s in wb_b.sheetnames]

def diff_sheet(xlsx_a: bytes, xlsx_b: bytes, sheet_name: str, max_changes: int = 5000) -> pd.DataFrame:
    wb_a = openpyxl.load_workbook(BytesIO(xlsx_a), data_only=True)
    wb_b = openpyxl.load_workbook(BytesIO(xlsx_b), data_only=True)

    if sheet_name not in wb_a.sheetnames or sheet_name not in wb_b.sheetnames:
        return pd.DataFrame(columns=["Cell", "Row", "Col", "A", "B"])

    ws_a = wb_a[sheet_name]
    ws_b = wb_b[sheet_name]

    max_row = max(ws_a.max_row or 1, ws_b.max_row or 1)
    max_col = max(ws_a.max_column or 1, ws_b.max_column or 1)

    changes: list[tuple[str,int,int,str,str]] = []
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            va = _norm(ws_a.cell(r, c).value)
            vb = _norm(ws_b.cell(r, c).value)
            if va != vb:
                changes.append((f"{get_column_letter(c)}{r}", r, c, va, vb))
                if len(changes) >= max_changes:
                    break
        if len(changes) >= max_changes:
            break

    return pd.DataFrame(changes, columns=["Cell", "Row", "Col", "A", "B"])

def diff_summary(diffs: pd.DataFrame) -> dict:
    if diffs.empty:
        return {"changed_cells": 0, "changed_rows": 0, "changed_cols": 0}
    return {
        "changed_cells": int(len(diffs)),
        "changed_rows": int(diffs["Row"].nunique()),
        "changed_cols": int(diffs["Col"].nunique()),
    }

def top_changed_rows(diffs: pd.DataFrame, top_n: int = 25) -> pd.DataFrame:
    if diffs.empty:
        return pd.DataFrame(columns=["Row", "ChangedCells"])
    return (
        diffs.groupby("Row", as_index=False)
        .size()
        .rename(columns={"size": "ChangedCells"})
        .sort_values("ChangedCells", ascending=False)
        .head(top_n)
    )

def top_changed_cols(diffs: pd.DataFrame, top_n: int = 25) -> pd.DataFrame:
    if diffs.empty:
        return pd.DataFrame(columns=["Col", "ChangedCells"])
    return (
        diffs.groupby("Col", as_index=False)
        .size()
        .rename(columns={"size": "ChangedCells"})
        .sort_values("ChangedCells", ascending=False)
        .head(top_n)
    )
